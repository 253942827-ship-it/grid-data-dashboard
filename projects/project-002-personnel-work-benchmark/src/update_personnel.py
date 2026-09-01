#!/usr/bin/env python3
"""从人员底层数据模板 Excel 更新 personnel.json / tangxia_personnel.json。"""
import os
import sys
import json
import openpyxl
from datetime import datetime

PROJ_DIR = "/Users/mr.g/Documents/Codex/Workspace/projects/project-002-personnel-work-benchmark"
ALLOWED_ROLES = {"装维经理", "片区经理", "营业员"}
REQUIRED = ["姓名", "揽装编码", "岗位"]


def num(v, default=0.0):
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def read_sheet(ws):
    headers = [str(c.value or "").strip() for c in ws[1]]
    missing = [h for h in REQUIRED if h not in headers]
    if missing:
        raise ValueError(f"Sheet {ws.title} 缺少列: {missing}")
    people = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        name = str(rec.get("姓名") or "").strip()
        code = str(rec.get("揽装编码") or "").strip()
        role = str(rec.get("岗位") or "").strip()
        if not name and not code:
            continue
        if not name or not code:
            raise ValueError(f"{ws.title} 存在姓名或编码为空的记录: 姓名={name} 编码={code}")
        if role not in ALLOWED_ROLES:
            raise ValueError(f"{ws.title} 岗位不合法: {name} -> {role}（只能为 {'/'.join(sorted(ALLOWED_ROLES))}）")
        if code in seen:
            raise ValueError(f"{ws.title} 揽装编码重复: {code} {name}")
        seen.add(code)
        people.append({
            "name": name,
            "code": code,
            "role": role,
            "cp": str(rec.get("CP") or "").strip(),
            "target_value": num(rec.get("价值积分目标")),
            "target_total": num(rec.get("增存高套目标")),
            "target_daliang": num(rec.get("达量目标")),
        })
    if not people:
        raise ValueError(f"{ws.title} 没有有效人员数据")
    return people


def main():
    if len(sys.argv) < 2:
        print("用法: python3 src/update_personnel.py <模板Excel路径>")
        return 1
    fp = sys.argv[1]
    wb = openpyxl.load_workbook(fp, data_only=True)
    satai = read_sheet(wb["沙太人员"])
    tangxia = read_sheet(wb["棠下人员"])
    today = datetime.now().strftime("%Y-%m-%d")
    for fn, people in [("personnel.json", satai), ("tangxia_personnel.json", tangxia)]:
        path = os.path.join(PROJ_DIR, "data", fn)
        with open(path, encoding="utf-8") as f:
            old = json.load(f)
        old["data_date"] = old.get("data_date") or today
        old["personnel"] = people
        with open(path, "w", encoding="utf-8") as f:
            json.dump(old, f, ensure_ascii=False, indent=2)
    print(f"✅ 已更新 personnel.json（沙太 {len(satai)} 人）和 tangxia_personnel.json（棠下 {len(tangxia)} 人）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
