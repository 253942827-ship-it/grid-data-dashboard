#!/usr/bin/env python3
"""个人视图生成脚本：搜索姓名，展示个人各项目进度，质态异常可下钻接入号。"""
import calendar
import json
import os
import openpyxl
from datetime import datetime, date
from collections import defaultdict

PROJ_DIR = "/Users/mr.g/Documents/Codex/Workspace/projects/project-002-personnel-work-benchmark"
DATA_DIR = os.path.join(PROJ_DIR, "data")
OUTPUT_DIR = os.path.join(PROJ_DIR, "docs")


def safe_float(v):
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ('', '-', '/', '#n/a', 'null', '<null>', 'none'):
        return None
    try:
        return float(v)
    except ValueError:
        return None


def _parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, int) and 19000000 <= v <= 21000000:
        try:
            return datetime.strptime(str(v), '%Y%m%d').date()
        except ValueError:
            return None
    if isinstance(v, str) and len(v) >= 10:
        for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S'):
            try:
                return datetime.strptime(v[:19] if fmt == '%Y-%m-%d %H:%M:%S' else v[:10], fmt).date()
            except ValueError:
                continue
    return None


def _list_month(fp):
    """返回清单文件报表日期列的最新月份（YYYY-MM），无日期或文件缺失返回 None。"""
    if not os.path.exists(fp):
        return None
    best = None
    try:
        wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
        for ws in wb.worksheets[:3]:
            date_cols = set()
            for r in range(1, 4):
                for c in range(1, min(ws.max_column, 60) + 1):
                    h = str(ws.cell(r, c).value or '').strip()
                    if h in ('录入时间', '统计日期', '竣工日期'):
                        date_cols.add(c)
            if not date_cols:
                continue
            for c in date_cols:
                for row in ws.iter_rows(min_row=2, max_row=30, min_col=c, max_col=c, values_only=True):
                    v = row[0]
                    m = None
                    if isinstance(v, (datetime, date)):
                        m = v.strftime('%Y-%m')
                    elif isinstance(v, int) and 20250101 <= v <= 20261231:
                        m = f"{v // 10000:04d}-{(v // 100) % 100:02d}"
                    elif isinstance(v, str):
                        s = v.strip()
                        if len(s) >= 7 and s[4] == '-' and s[:4].isdigit():
                            m = s[:7]
                        elif len(s) == 8 and s.isdigit() and 20250101 <= int(s) <= 20261231:
                            m = f"{int(s) // 10000:04d}-{(int(s) // 100) % 100:02d}"
                    if m and (not best or m > best):
                        best = m
        wb.close()
    except Exception:
        return None
    return best


def _infer_data_month():
    fp = os.path.join(DATA_DIR, "新装高套竣工清单.xlsx")
    if not os.path.exists(fp):
        return datetime.now().strftime('%Y-%m')
    wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
    latest = None
    try:
        ws = wb.active
        for row in ws.iter_rows(min_row=2, min_col=8, max_col=8, values_only=True):
            d = _parse_date(row[0])
            if d and d >= date(2026, 1, 1) and (not latest or d > latest):
                latest = d
    finally:
        wb.close()
    return latest.strftime('%Y-%m') if latest else datetime.now().strftime('%Y-%m')


def _header_cols(ws, aliases, rows=(1, 2, 3)):
    out = {}
    for r in rows:
        for c in range(1, ws.max_column + 1):
            key = aliases.get(str(ws.cell(r, c).value or '').strip())
            if key:
                out.setdefault(key, c)
    return out


def main():
    pd_json = json.load(open(os.path.join(DATA_DIR, "personnel.json"), encoding="utf-8"))
    personnel = pd_json["personnel"]
    target_month = str(pd_json.get("target_month", "") or "")
    by_code = {str(p.get("code", "")).strip(): p for p in personnel}
    names = {p["name"] for p in personnel}

    benchmark = _infer_data_month()
    bench_y, bench_m = int(benchmark[:4]), int(benchmark[5:7])
    last_key = f"{bench_y}-{bench_m - 1:02d}" if bench_m > 1 else f"{bench_y - 1}-12"
    last_y, last_m = int(last_key[:4]), int(last_key[5:7])
    t0_key = benchmark.replace("-", "")
    t1_key = last_key.replace("-", "")
    cal_days = calendar.monthrange(bench_y, bench_m)[1]
    last_cal_days = calendar.monthrange(last_y, last_m)[1]
    ok = lambda f: _list_month(os.path.join(DATA_DIR, f)) == benchmark
    exist_ok = ok("存量高套竣工清单.xlsx")
    gb_ok = ok("杠保清单.xlsx")
    ko_ok = ok("关键一单清单.xlsx")
    zt_ok = ok("质态相关清单.xlsx")
    print(f"  个人视图基准月份: {benchmark} 存量={exist_ok} 杠保={gb_ok} 关键一单={ko_ok} 质态={zt_ok}")

    # ---- 新装 / 存量 / 上月 积分与高套 ----
    new_install = defaultdict(lambda: {"score": 0.0, "gaotao": 0.0})
    exist_install = defaultdict(lambda: {"score": 0.0, "gaotao": 0.0})
    last_new = defaultdict(float)
    last_exist = defaultdict(float)
    new_accs = defaultdict(list)
    exist_accs = defaultdict(list)
    cur_dates, last_dates = [], []

    fp = os.path.join(DATA_DIR, "新装高套竣工清单.xlsx")
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    cols = _header_cols(ws, {"揽装人": "name", "套餐价值": "value", "折算后": "gaotao", "竣工日期": "date", "接入号": "acc"})
    cn, cv, cg, cd, ca = cols.get("name", 11), cols.get("value", 15), cols.get("gaotao", 26), cols.get("date", 8), cols.get("acc", 4)
    tech = any(str(ws.cell(2, c).value or "").strip() == "sales_name" for c in range(1, ws.max_column + 1))
    for r in range(3 if tech else 2, ws.max_row + 1):
        name = str(ws.cell(r, cn).value or "").strip()
        if not name:
            continue
        d = _parse_date(ws.cell(r, cd).value)
        if d is not None and d.strftime("%Y-%m") != benchmark:
            continue
        new_install[name]["score"] += safe_float(ws.cell(r, cv).value) or 0
        new_install[name]["gaotao"] += safe_float(ws.cell(r, cg).value) or 0
        acc = str(ws.cell(r, ca).value or "").strip()
        if acc:
            new_accs[name].append(acc)
        if d:
            cur_dates.append(d)
    wb.close()

    if exist_ok:
        fp = os.path.join(DATA_DIR, "存量高套竣工清单.xlsx")
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb.active
        cols = _header_cols(ws, {
            "揽装人": "name", "提值幅度": "value", "高套折算量": "gaotao", "竣工日期": "date",
            "sj_salestaff_name": "name", "jzbh_value": "value", "gt_zsl": "gaotao", "sj_subs_stat_date": "date",
            "接入号": "acc",
        })
        cn, cv, cg, cd, ca = cols.get("name", 10), cols.get("value", 16), cols.get("gaotao", 27), cols.get("date", 11), cols.get("acc", 6)
        tech = any(str(ws.cell(2, c).value or "").strip() == "sj_salestaff_name" for c in range(1, ws.max_column + 1))
        for r in range(3 if tech else 2, ws.max_row + 1):
            name = str(ws.cell(r, cn).value or "").strip()
            if not name:
                continue
            d = _parse_date(ws.cell(r, cd).value)
            if d is not None and d.strftime("%Y-%m") != benchmark:
                continue
            exist_install[name]["score"] += safe_float(ws.cell(r, cv).value) or 0
            exist_install[name]["gaotao"] += safe_float(ws.cell(r, cg).value) or 0
            acc = str(ws.cell(r, ca).value or "").strip()
            if acc:
                exist_accs[name].append(acc)
            if d:
                cur_dates.append(d)
        wb.close()

    fp = os.path.join(DATA_DIR, "上月新装高套清单.xlsx")
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    for r in range(3, ws.max_row + 1):
        name = str(ws.cell(r, 11).value or "").strip()
        if not name:
            continue
        d = _parse_date(ws.cell(r, 8).value)
        if d is not None and d.strftime("%Y-%m") != last_key:
            continue
        last_new[name] += safe_float(ws.cell(r, 15).value) or 0
        if d:
            last_dates.append(d)
    wb.close()

    fp = os.path.join(DATA_DIR, "上月存量高套清单.xlsx")
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    cols = _header_cols(ws, {"揽装人": "name", "提值幅度": "value", "竣工日期": "date"})
    cn, cv, cd = cols.get("name", 14), cols.get("value", 21), cols.get("date", 15)
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, cn).value or "").strip()
        if not name:
            continue
        d = _parse_date(ws.cell(r, cd).value)
        if d is not None and d.strftime("%Y-%m") != last_key:
            continue
        last_exist[name] += safe_float(ws.cell(r, cv).value) or 0
        if d:
            last_dates.append(d)
    wb.close()

    cur_days = (max(cur_dates) - min(cur_dates)).days + 1 if cur_dates else cal_days
    last_days = (max(last_dates) - min(last_dates)).days + 1 if last_dates else last_cal_days

    # ---- 杠保 ----
    gb = defaultdict(lambda: {"total": 0, "success": 0})
    if gb_ok:
        fp = os.path.join(DATA_DIR, "杠保清单.xlsx")
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb.active
        for r in range(3, ws.max_row + 1):
            name = str(ws.cell(r, 8).value or "").strip()
            if not name or name == "sales_name":
                continue
            d = _parse_date(ws.cell(r, 4).value)
            if d is not None and d.strftime("%Y-%m") != benchmark:
                continue
            is_gb = safe_float(ws.cell(r, 20).value) or 0
            gb[name]["total"] += 1
            if is_gb == 1:
                gb[name]["success"] += 1
        wb.close()

    # ---- 关键一单 ----
    ko = defaultdict(lambda: {"dispatch": 0, "convert": 0})
    if ko_ok:
        fp = os.path.join(DATA_DIR, "关键一单清单.xlsx")
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb.active
        cols = _header_cols(ws, {
            "施工人姓名": "name", "施工人工号": "code", "订单状态名称": "state",
            "实时受理积分": "real", "竣工积分": "finish",
        })
        cn, cc = cols.get("name") or cols.get("code", 26), cols.get("code", 26)
        cs, cr, cf = cols.get("state", 3), cols.get("real", 101), cols.get("finish", 110)
        for r in range(2, ws.max_row + 1):
            name = str(ws.cell(r, cn).value or "").strip()
            if not name:
                continue
            d = _parse_date(ws.cell(r, 2).value)
            if d is not None and d.strftime("%Y-%m") != benchmark:
                continue
            if str(ws.cell(r, cs).value or "").strip() == "作废":
                continue
            ko[name]["dispatch"] += 1
            if (safe_float(ws.cell(r, cr).value) or 0) > 0 or (safe_float(ws.cell(r, cf).value) or 0) > 0:
                ko[name]["convert"] += 1
        wb.close()

    # ---- 质态异常（含接入号明细） ----
    zt = {}
    if zt_ok:
        fp = os.path.join(DATA_DIR, "质态相关清单.xlsx")
        wb = openpyxl.load_workbook(fp, data_only=True)
        for sheet in wb.sheetnames:
            if sheet.startswith("融合质态T+0"):
                ws = wb[sheet]
                for r in range(3, ws.max_row + 1):
                    name = str(ws.cell(r, 10).value or "").strip()
                    if not name:
                        continue
                    month = str(ws.cell(r, 3).value or "").strip()
                    acc = str(ws.cell(r, 4).value or "").strip()
                    if month == t0_key:
                        if (safe_float(ws.cell(r, 5).value) or 0) == 0:
                            zt.setdefault(name, {"t0_invalid": [], "t0_notrust": [], "t1_invalid": [], "t1_notrust": [], "t0_notfull": []})["t0_invalid"].append(acc)
                        if (safe_float(ws.cell(r, 7).value) or 0) == 0:
                            zt.setdefault(name, {"t0_invalid": [], "t0_notrust": [], "t1_invalid": [], "t1_notrust": [], "t0_notfull": []})["t0_notrust"].append(acc)
                    elif month == t1_key:
                        if (safe_float(ws.cell(r, 5).value) or 0) == 0:
                            zt.setdefault(name, {"t0_invalid": [], "t0_notrust": [], "t1_invalid": [], "t1_notrust": [], "t0_notfull": []})["t1_invalid"].append(acc)
                        if (safe_float(ws.cell(r, 7).value) or 0) == 0:
                            zt.setdefault(name, {"t0_invalid": [], "t0_notrust": [], "t1_invalid": [], "t1_notrust": [], "t0_notfull": []})["t1_notrust"].append(acc)
            elif "未满卡" in sheet:
                ws = wb[sheet]
                for r in range(3, ws.max_row + 1):
                    name = str(ws.cell(r, 8).value or "").strip()
                    if not name or name == "sales_name":
                        continue
                    if str(ws.cell(r, 4).value or "").strip() in ("是", "1"):
                        continue
                    zt.setdefault(name, {"t0_invalid": [], "t0_notrust": [], "t1_invalid": [], "t1_notrust": [], "t0_notfull": []})["t0_notfull"].append(str(ws.cell(r, 3).value or "").strip())
        wb.close()
        for v in zt.values():
            for k in v:
                v[k] = list(dict.fromkeys(x for x in v[k] if x))

    # ---- 商客发展 ----
    sk = {}
    sk_path = os.path.join(DATA_DIR, "商客发展情况.json")
    sk_meta = None
    if os.path.exists(sk_path):
        try:
            sk_meta = json.load(open(sk_path, encoding="utf-8"))
            sk = sk_meta.get("people", {})
        except Exception:
            sk = {}

    # ---- 组装个人数据 ----
    people = {}
    for p in personnel:
        code = str(p.get("code", "")).strip()
        name = p["name"]
        ni = new_install.get(name, {})
        ei = exist_install.get(name, {})
        new_score = round(ni.get("score", 0), 2)
        exist_score = round(ei.get("score", 0), 2) if exist_ok else None
        total_score = round(new_score + (exist_score or 0), 2)
        last_total = round(last_new.get(name, 0) + last_exist.get(name, 0), 2)
        mom = None
        if last_total and last_total != 0:
            mom = round((total_score / cur_days - last_total / last_days) / abs(last_total / last_days), 4)
        new_gaotao = round(ni.get("gaotao", 0), 2)
        exist_gaotao = round(ei.get("gaotao", 0), 2) if exist_ok else None
        total_gaotao = round(new_gaotao + (exist_gaotao or 0), 2)
        target_value = float(p.get("target_value", 0) or 0)
        target_total = float(p.get("target_total", 0) or 0)
        target_daliang = float(p.get("target_daliang", 0) or 0)
        value_rate = round(total_score / target_value, 4) if target_value > 0 else None
        gaotao_rate = round(total_gaotao / target_total, 4) if target_total > 0 else None
        gap = round(target_daliang - total_gaotao, 2) if target_daliang > 0 else None
        g = gb.get(name, {})
        gb_rate = round(g.get("success", 0) / g["total"], 4) if gb_ok and g.get("total") else None
        k = ko.get(name, {})
        ko_rate = round(k.get("convert", 0) / k["dispatch"], 4) if ko_ok and k.get("dispatch") else None
        if p["role"] != "装维经理":
            ko_rate = None
        z = zt.get(name, {"t0_invalid": [], "t0_notrust": [], "t1_invalid": [], "t1_notrust": [], "t0_notfull": []}) if zt_ok else {"t0_invalid": [], "t0_notrust": [], "t1_invalid": [], "t1_notrust": [], "t0_notfull": []}
        s = sk.get(code, {})
        people[code] = {
            "name": name,
            "code": code,
            "role": p["role"],
            "cp": p.get("cp", "") or "",
            "target_value": target_value,
            "target_total": target_total,
            "target_daliang": target_daliang,
            "new_score": new_score,
            "exist_score": exist_score,
            "total_score": total_score,
            "last_total": last_total,
            "mom": mom,
            "value_rate": value_rate,
            "new_gaotao": new_gaotao,
            "exist_gaotao": exist_gaotao,
            "total_gaotao": total_gaotao,
            "new_accs": list(dict.fromkeys(new_accs.get(name, []))),
            "exist_accs": list(dict.fromkeys(exist_accs.get(name, []))),
            "gaotao_rate": gaotao_rate,
            "gap": gap,
            "gb_total": g.get("total", 0) if gb_ok else None,
            "gb_success": g.get("success", 0) if gb_ok else None,
            "gb_rate": gb_rate,
            "ko_dispatch": k.get("dispatch", 0) if ko_ok and p["role"] == "装维经理" else None,
            "ko_convert": k.get("convert", 0) if ko_ok and p["role"] == "装维经理" else None,
            "ko_rate": ko_rate,
            "zt": z,
            "sk_shangqi": s.get("shangqi"),
            "sk_weixiao": s.get("weixiao"),
            "sk_dev299": s.get("dev299"),
        }

    zw_rates = [people[c]["ko_rate"] for c in people if people[c]["ko_rate"] is not None]
    avg_ko_rate = round(sum(zw_rates) / len(zw_rates), 4) if zw_rates else None

    updated_at = sk_meta.get("updated_at", datetime.now().strftime("%Y-%m-%d %H:%M")) if sk_meta else datetime.now().strftime("%Y-%m-%d %H:%M")
    order = list(people.keys())
    payload = {
        "benchmark": benchmark,
        "target_month": target_month,
        "updated_at": updated_at,
        "avg_ko_rate": avg_ko_rate,
        "order": order,
        "people": people,
    }
    data_json = json.dumps(payload, ensure_ascii=False)

    html = _HTML_TPL.replace("__DATA__", data_json)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out = os.path.join(OUTPUT_DIR, "personal.html")
    with open(out, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"\n✅ 个人视图已生成: {out}")
    print(f"   文件大小: {os.path.getsize(out) / 1024:.0f} KB")


_HTML_TPL = r'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>个人视图 - 沙太人员工作看板</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,'PingFang SC','Microsoft YaHei',sans-serif;background:#eef1f6;color:#1a1a2e;padding:20px}
.page{max-width:1080px;margin:0 auto}
.topbar{background:linear-gradient(135deg,#0d1b4a,#1a237e);color:#fff;border-radius:10px;padding:16px 22px;margin-bottom:16px}
.topbar h1{font-size:18px;font-weight:700}
.topbar .sub{font-size:11px;opacity:.75;margin-top:4px}
.search-wrap{margin-top:12px;position:relative;max-width:420px}
.search-wrap input{width:100%;padding:10px 14px;border:none;border-radius:8px;font-size:14px;outline:none;font-family:inherit}
.dropdown{position:absolute;top:calc(100% + 4px);left:0;right:0;background:#fff;border-radius:8px;box-shadow:0 6px 20px rgba(0,0,0,.15);max-height:280px;overflow:auto;z-index:10;display:none}
.dropdown .item{padding:9px 14px;cursor:pointer;font-size:13px;border-bottom:1px solid #f0f0f0}
.dropdown .item:hover,.dropdown .item.active{background:#e8eaf6;color:#1a237e}
.card{background:#fff;border-radius:10px;box-shadow:0 1px 6px rgba(0,0,0,.05);padding:14px 16px;margin-bottom:14px}
.person-head{display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px}
.person-head .name{font-size:20px;font-weight:700;color:#0d1b4a}
.person-head .role{font-size:12px;color:#888}
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px;margin-bottom:14px}
.kpi{background:#fff;border-radius:10px;padding:12px 14px;box-shadow:0 1px 6px rgba(0,0,0,.05)}
.kpi .num{font-size:22px;font-weight:700;font-variant-numeric:tabular-nums}
.kpi .label{font-size:11px;color:#888;margin-top:2px}
.bar{height:7px;background:#e8eaf6;border-radius:4px;overflow:hidden;margin-top:7px}
.bar i{display:block;height:100%;border-radius:4px}
.sec-title{font-size:13px;font-weight:700;color:#1a237e;margin-bottom:10px;border-bottom:2px solid #e8eaf6;padding-bottom:6px}
.grid2{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:10px}
.stat{background:#f8f9ff;border:1px solid #eef0f7;border-radius:8px;padding:10px 12px}
.stat .lab{font-size:11px;color:#888}
.stat .val{font-size:17px;font-weight:600;margin-top:3px;font-variant-numeric:tabular-nums}
.stat .val.small{font-size:14px}
.positive{color:#2e7d32}.negative{color:#c62828}.warn{color:#e65100}
.click{color:#1a237e;cursor:pointer;text-decoration:underline}
.click:hover{color:#c62828}
.na{color:#ccc}
.modal-mask{position:fixed;inset:0;background:rgba(13,27,74,.45);display:none;align-items:center;justify-content:center;z-index:99;padding:20px}
.modal{background:#fff;border-radius:10px;max-width:720px;width:100%;max-height:80vh;display:flex;flex-direction:column}
.modal-head{padding:12px 16px;border-bottom:2px solid #e8eaf6;display:flex;justify-content:space-between;align-items:center}
.modal-head b{color:#1a237e}
.modal-close{border:none;background:#e8eaf6;color:#1a237e;border-radius:6px;padding:4px 10px;cursor:pointer;font-size:13px}
.modal-body{overflow:auto;padding:14px 16px}
.modal-body table{width:100%;border-collapse:collapse;font-size:12px}
.modal-body th{background:#e8eaf6;color:#1a237e;padding:6px;text-align:left;position:sticky;top:0}
.modal-body td{padding:5px 6px;border-bottom:1px solid #f0f0f0;font-variant-numeric:tabular-nums}
.empty{color:#999;font-size:12px;padding:6px 0}
.footer{text-align:center;font-size:11px;color:#aaa;margin-top:10px}
</style>
</head>
<body>
<div class="page">
  <div class="topbar">
    <h1>👤 沙太人员个人视图</h1>
    <div class="sub">数据月份 <span id="benchmark"></span> · 目标月份 <span id="targetMonth"></span> · 更新时间 <span id="updatedAt"></span></div>
    <div class="search-wrap">
      <input id="search" placeholder="输入姓名搜索" autocomplete="off">
      <div class="dropdown" id="dropdown"></div>
    </div>
  </div>

  <div id="personView" style="display:none">
    <div class="card person-head">
      <div>
        <span class="name" id="pName"></span>
        <span class="role" id="pRole"></span>
      </div>
      <div style="font-size:12px;color:#888">揽装编码 <span id="pCode"></span> · CP <span id="pCp"></span></div>
    </div>

    <div class="kpis">
      <div class="kpi"><div class="num" id="kScore"></div><div class="label">价值积分 · 目标完成</div><div class="bar"><i id="kScoreBar"></i></div></div>
      <div class="kpi"><div class="num" id="kGaotao"></div><div class="label">增存高套 · 目标完成</div><div class="bar"><i id="kGaotaoBar"></i></div></div>
      <div class="kpi"><div class="num" id="kGap"></div><div class="label">达量缺口</div></div>
    </div>

    <div class="card">
      <div class="sec-title">① 价值积分</div>
      <div class="grid2">
        <div class="stat"><div class="lab">新装积分</div><div class="val" id="sNew"></div></div>
        <div class="stat"><div class="lab">存量积分</div><div class="val" id="sExist"></div></div>
        <div class="stat"><div class="lab">合计 / 目标</div><div class="val" id="sTotal"></div></div>
        <div class="stat"><div class="lab">环比（日均）</div><div class="val" id="sMom"></div></div>
      </div>
    </div>

    <div class="card">
      <div class="sec-title">② 增存高套（点击折算数查看接入号）</div>
      <div class="grid2">
        <div class="stat"><div class="lab">新装折算（点击查看接入号）</div><div class="val" id="gNew"></div></div>
        <div class="stat"><div class="lab">存量折算（点击查看接入号）</div><div class="val" id="gExist"></div></div>
        <div class="stat"><div class="lab">总计 / 目标</div><div class="val" id="gTotal"></div></div>
        <div class="stat"><div class="lab">达量目标 / 缺口</div><div class="val" id="gGap"></div></div>
      </div>
    </div>

    <div class="card">
      <div class="sec-title">③ 杠保</div>
      <div class="grid2">
        <div class="stat"><div class="lab">杠保成功量</div><div class="val" id="gbSuccess"></div></div>
        <div class="stat"><div class="lab">成功率</div><div class="val" id="gbRate"></div></div>
      </div>
    </div>

    <div class="card">
      <div class="sec-title">④ 关键一单（装维经理）</div>
      <div class="grid2">
        <div class="stat"><div class="lab">派单 / 转化</div><div class="val" id="koCount"></div></div>
        <div class="stat"><div class="lab">转化率 vs 营服平均</div><div class="val" id="koRate"></div></div>
      </div>
    </div>

    <div class="card">
      <div class="sec-title">⑥ 质态异常（点击数字查看异常接入号）</div>
      <div class="grid2" id="ztList"></div>
    </div>

    <div class="card">
      <div class="sec-title">⑦ 商客发展</div>
      <div class="grid2">
        <div class="stat"><div class="lab">增存商企</div><div class="val" id="skShangqi"></div></div>
        <div class="stat"><div class="lab">小微 ICT</div><div class="val" id="skWeixiao"></div></div>
        <div class="stat"><div class="lab">299+ 发展量</div><div class="val" id="skDev299"></div></div>
      </div>
    </div>
  </div>
  <div class="footer">沙太人员个人视图 · 数据来自各月度清单，仅供工作核对</div>
</div>

<div class="modal-mask" id="modal">
  <div class="modal">
    <div class="modal-head"><b id="modalTitle"></b><button class="modal-close" onclick="closeModal()">关闭</button></div>
    <div class="modal-body"><table><thead><tr><th>接入号</th></tr></thead><tbody id="modalBody"></tbody></table></div>
  </div>
</div>

<script>
var DATA = __DATA__;
var current = null;
var ZT_LABEL={'t0_notfull':'T0 未满卡','t0_invalid':'T0 无效','t0_notrust':'T0 无托收','t1_invalid':'T1 无效','t1_notrust':'T1 无托收'};

function esc(s){return String(s==null?'':s).replace(/[&<>"]/g,function(c){return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]})}
function fmt(v, d){if(v===null||v===undefined||v===''||isNaN(v))return '<span class="na">-</span>';return Number(v).toFixed(d==null?1:d)}
function pct(v){if(v===null||v===undefined||isNaN(v))return '<span class="na">-</span>';return (v*100).toFixed(1)+'%'}
function momCls(v){if(v===null||v===undefined||isNaN(v))return 'na';return v>=0?'positive':'negative'}
function rateCls(v){if(v===null||v===undefined||isNaN(v))return '';return v>=0.8?'positive':(v>=0.5?'warn':'negative')}

function fillBar(id, rate){
  var el=document.getElementById(id);
  if(rate===null||rate===undefined||isNaN(rate)){el.style.width='0';el.style.background='#ccc';return}
  var w=Math.min(100,rate*100);
  el.style.width=w+'%';
  el.style.background=rate>=0.8?'#2e7d32':(rate>=0.5?'#e65100':'#c62828');
}

function render(code){
  var p=DATA.people[code]; if(!p)return;
  current=code;
  document.getElementById('personView').style.display='block';
  document.getElementById('pName').textContent=p.name;
  document.getElementById('pRole').textContent=p.role;
  document.getElementById('pCode').textContent=p.code;
  document.getElementById('pCp').textContent=p.cp||'-';

  document.getElementById('kScore').innerHTML=fmt(p.total_score,1)+' / '+fmt(p.target_value,1);
  fillBar('kScoreBar', p.value_rate);
  document.getElementById('kGaotao').innerHTML=fmt(p.total_gaotao,1)+' / '+fmt(p.target_total,1);
  fillBar('kGaotaoBar', p.gaotao_rate);
  document.getElementById('kGap').innerHTML=p.gap===null?'<span class="na">-</span>':(p.gap>=0?fmt(p.gap,1):'<span class="negative">'+fmt(p.gap,1)+'</span>');

  document.getElementById('sNew').innerHTML=fmt(p.new_score,1);
  document.getElementById('sExist').innerHTML=fmt(p.exist_score,1);
  document.getElementById('sTotal').innerHTML=fmt(p.total_score,1)+' / '+fmt(p.target_value,1);
  var mom=p.mom;
  document.getElementById('sMom').innerHTML=mom===null?'<span class="na">-</span>':'<span class="'+momCls(mom)+'">'+(mom>=0?'+':'')+(mom*100).toFixed(1)+'%</span>';

  document.getElementById('gNew').innerHTML=(p.new_accs&&p.new_accs.length)?'<span class="click" onclick="showAcc(\''+code+'\',\'new_accs\',\'新装接入号\')">'+fmt(p.new_gaotao,1)+'</span>':fmt(p.new_gaotao,1);
  document.getElementById('gExist').innerHTML=(p.exist_accs&&p.exist_accs.length)?'<span class="click" onclick="showAcc(\''+code+'\',\'exist_accs\',\'存量接入号\')">'+fmt(p.exist_gaotao,1)+'</span>':fmt(p.exist_gaotao,1);
  document.getElementById('gTotal').innerHTML=fmt(p.total_gaotao,1)+' / '+fmt(p.target_total,1);
  document.getElementById('gGap').innerHTML=p.gap===null?'<span class="na">-</span>':fmt(p.gap,1);

  document.getElementById('gbSuccess').innerHTML=fmt(p.gb_success,0);
  document.getElementById('gbRate').innerHTML=pct(p.gb_rate);

  document.getElementById('koCount').innerHTML=p.ko_dispatch===null?'<span class="na">-</span>':fmt(p.ko_dispatch,0)+' / '+fmt(p.ko_convert,0);
  var kr=p.ko_rate, avg=DATA.avg_ko_rate;
  document.getElementById('koRate').innerHTML=(kr===null?'<span class="na">-</span>':pct(kr))+' <span style="font-size:11px;color:#888">vs 营服均值 '+pct(avg)+'</span>';

  var zt=p.zt, ztHtml='';
  var ztItems=[['t0_notfull','T0 未满卡'],['t0_invalid','T0 无效'],['t0_notrust','T0 无托收'],['t1_invalid','T1 无效'],['t1_notrust','T1 无托收']];
  ztItems.forEach(function(it){
    var arr=zt[it[0]]||[]; var n=arr.length;
    ztHtml+='<div class="stat"><div class="lab">'+it[1]+'</div><div class="val '+(n>0?'warn':'')+'">'+
      (n>0?'<span class="click" onclick="showZt(\''+code+'\',\''+it[0]+'\')">'+n+' 条</span>':'<span class="na">0</span>')+'</div></div>';
  });
  document.getElementById('ztList').innerHTML=ztHtml;

  document.getElementById('skShangqi').innerHTML=fmt(p.sk_shangqi,0);
  document.getElementById('skWeixiao').innerHTML=fmt(p.sk_weixiao,1);
  document.getElementById('skDev299').innerHTML=fmt(p.sk_dev299,0);
}

function showZt(code, key){
  var arr=DATA.people[code].zt[key]||[];
  document.getElementById('modalTitle').textContent=DATA.people[code].name+' · '+(ZT_LABEL[key]||key)+' 异常接入号（'+arr.length+' 条）';
  var body=document.getElementById('modalBody');
  body.innerHTML=arr.map(function(x){return '<tr><td>'+esc(x)+'</td></tr>'}).join('')||'<tr><td class="empty">无</td></tr>';
  document.getElementById('modal').style.display='flex';
}
function showAcc(code, key, label){
  var arr=DATA.people[code][key]||[];
  document.getElementById('modalTitle').textContent=DATA.people[code].name+' · '+label+'（'+arr.length+' 条）';
  var body=document.getElementById('modalBody');
  body.innerHTML=arr.map(function(x){return '<tr><td>'+esc(x)+'</td></tr>'}).join('')||'<tr><td class="empty">无</td></tr>';
  document.getElementById('modal').style.display='flex';
}
function closeModal(){document.getElementById('modal').style.display='none'}
document.getElementById('modal').addEventListener('click',function(e){if(e.target===this)closeModal()});

var search=document.getElementById('search'), dd=document.getElementById('dropdown');
var filtered=[], activeIdx=-1;
function updateDrop(){
  var q=search.value.trim();
  if(!q){dd.style.display='none';return}
  filtered=DATA.order.map(function(c){return DATA.people[c]}).filter(function(p){return p.name.indexOf(q)>=0||p.code.indexOf(q)>=0});
  if(!filtered.length){dd.style.display='none';return}
  activeIdx=0;
  dd.innerHTML=filtered.map(function(p,i){return '<div class="item'+(i===0?' active':'')+'" data-code="'+p.code+'">'+esc(p.name)+' <span style="color:#aaa;font-size:11px">'+esc(p.role)+'</span></div>'}).join('');
  dd.style.display='block';
}
function pick(code){
  search.value=DATA.people[code].name;
  dd.style.display='none';
  render(code);
}
search.addEventListener('input',updateDrop);
search.addEventListener('keydown',function(e){
  if(e.key==='Enter'){if(filtered[activeIdx])pick(filtered[activeIdx].code)}
  else if(e.key==='ArrowDown'&&filtered.length){activeIdx=(activeIdx+1)%filtered.length;mark()}
  else if(e.key==='ArrowUp'&&filtered.length){activeIdx=(activeIdx-1+filtered.length)%filtered.length;mark()}
});
dd.addEventListener('click',function(e){var it=e.target.closest('.item');if(it)pick(it.getAttribute('data-code'))});
function mark(){Array.prototype.forEach.call(dd.children,function(el,i){el.className='item'+(i===activeIdx?' active':'')})}
document.addEventListener('click',function(e){if(!search.contains(e.target)&&!dd.contains(e.target))dd.style.display='none'});

document.getElementById('benchmark').textContent=DATA.benchmark;
document.getElementById('targetMonth').textContent=DATA.target_month||'-';
document.getElementById('updatedAt').textContent=DATA.updated_at;
var urlName=new URLSearchParams(location.search).get('name');
var hit=urlName?DATA.order.map(function(c){return DATA.people[c]}).find(function(p){return p.name===urlName}):null;
if(hit){search.value=hit.name;render(hit.code)}
</script>
</body>
</html>
'''


if __name__ == "__main__":
    main()
