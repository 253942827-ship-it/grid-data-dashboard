#!/usr/bin/env python3
"""
人员工作数据对标看板生成脚本
读取多个外部清单文件，对标30位员工的6大维度指标，输出Web看板
"""

import openpyxl, json, os, math
from datetime import datetime, date
from collections import defaultdict

PROJ_DIR = "/Users/mr.g/Documents/Codex/Workspace/projects/project-002-personnel-work-benchmark"
DATA_DIR = os.path.join(PROJ_DIR, "data")
OUTPUT_DIR = os.path.join(PROJ_DIR, "docs")

# ============================================================
# 1. 读取人员基础信息 + 目标值（从主文件 Sheet1）
# ============================================================
print("1. 读取人员基础信息...")
import json
_pf = os.path.join(DATA_DIR, "personnel.json")
with open(_pf, 'r', encoding='utf-8') as _f:
    _pd = json.load(_f)
personnel = _pd['personnel']

def safe_float(v):
    if v is None: return None
    s = str(v).strip().lower()
    if s in ('', '-', '/', '#n/a', 'null', '<null>', 'none'):
        return None
    try: return float(v)
    except: return None

def _infer_data_month():
    """Use the latest date in the new-install list instead of the run date."""
    fp = os.path.join(DATA_DIR, "新装高套竣工清单.xlsx")
    if not os.path.exists(fp):
        return datetime.now().strftime('%Y-%m')
    wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
    latest = None
    try:
        ws = wb.active
        for row in ws.iter_rows(min_row=2, min_col=8, max_col=8, values_only=True):
            v = row[0]
            d = None
            if isinstance(v, (date, datetime)):
                d = v if isinstance(v, date) else v.date()
            elif isinstance(v, int) and 19000000 <= v <= 21000000:
                try:
                    d = datetime.strptime(str(v), '%Y%m%d').date()
                except ValueError:
                    pass
            elif isinstance(v, str) and len(v) >= 10:
                for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S'):
                    try:
                        d = datetime.strptime(v[:19], fmt).date()
                        break
                    except ValueError:
                        pass
            if isinstance(d, datetime):
                d = d.date()
            if d and d >= date(2026, 1, 1) and (not latest or d > latest):
                latest = d
    finally:
        wb.close()
    return latest.strftime('%Y-%m') if latest else datetime.now().strftime('%Y-%m')

data_date = _infer_data_month()

def _header_cols(ws, aliases, rows=(1, 2, 3)):
    """Locate Chinese/technical header columns in the first few rows."""
    out = {}
    for r in rows:
        for c in range(1, ws.max_column + 1):
            key = aliases.get(str(ws.cell(r, c).value or '').strip())
            if key:
                out.setdefault(key, c)
    return out

print(f"  共 {len(personnel)} 人")
role_order = {'装维经理': 0, '片区经理': 1, '营业员': 2}
personnel.sort(key=lambda p: (role_order.get(p['role'], 9), p['name']))

# ============================================================
# 1.5. 商客发展情况（全渠道做商客战报 · 人员维度）
# ============================================================
_shangke = {}
_shangke_meta = None
_fp_sk = os.path.join(DATA_DIR, "商客发展情况.json")
if os.path.exists(_fp_sk):
    try:
        with open(_fp_sk, "r", encoding="utf-8") as _f:
            _sk_data = json.load(_f)
        _shangke = _sk_data.get("people", {})
        _shangke_meta = _sk_data
    except Exception:
        _shangke = {}
for _p in personnel:
    _rec = _shangke.get(str(_p.get("code", "")).strip())
    _p["sk_shangqi"] = _rec["shangqi"] if _rec else None
    _p["sk_weixiao"] = _rec["weixiao"] if _rec else None
    _p["sk_dev299"] = _rec["dev299"] if _rec else None

# ============================================================
# 2. 新装高套竣工清单 → 新装价值积分 + 新装高套折算
# ============================================================
print("2. 读取新装高套竣工清单...")
new_install = defaultdict(lambda: {'value_score': 0, 'gaotao': 0})
fp = os.path.join(DATA_DIR, "新装高套竣工清单.xlsx")
wb2 = openpyxl.load_workbook(fp, data_only=True)
ws2 = wb2.active
cols2 = _header_cols(ws2, {
    '揽装人': 'name', '套餐价值': 'value', '折算后': 'gaotao', '竣工日期': 'date',
})
col_name2 = cols2.get('name', 11)
col_value2 = cols2.get('value', 15)
col_gaotao2 = cols2.get('gaotao', 26)
col_date2 = cols2.get('date', 8)
_new_has_technical_row = any(
    str(ws2.cell(2, c).value or '').strip() == 'sales_name'
    for c in range(1, ws2.max_column + 1)
)
_current_dates_new = []
for r in range(3 if _new_has_technical_row else 2, ws2.max_row + 1):
    name = str(ws2.cell(r, col_name2).value or '').strip()
    if not name: continue
    pv = safe_float(ws2.cell(r, col_value2).value) or 0  # 套餐价值
    zh = safe_float(ws2.cell(r, col_gaotao2).value) or 0    # 折算后
    new_install[name]['value_score'] += pv
    new_install[name]['gaotao'] += zh
    # 收集竣工日期
    dv = ws2.cell(r, col_date2).value
    if isinstance(dv, (datetime, date)):
        _current_dates_new.append(dv if isinstance(dv, date) else dv.date())
wb2.close()

# ============================================================
# 3. 存量高套竣工清单 → 存量价值积分 + 存量高套折算
# ============================================================


# ============================================================
# 4. 上月新装/存量（独立文件）→ 上月价值积分
# ============================================================

# ============================================================
# 3. 存量高套竣工清单 → 存量价值积分 + 存量高套折算
# ============================================================
print("3. 读取存量高套竣工清单...")
exist_install = defaultdict(lambda: {'value_score': 0, 'gaotao': 0})
fp = os.path.join(DATA_DIR, "存量高套竣工清单.xlsx")
wb3 = openpyxl.load_workbook(fp, data_only=True)
ws3 = wb3.active
# 动态检测列号
cols3 = _header_cols(ws3, {
    '揽装人': 'name', '提值幅度': 'value', '高套折算量': 'gaotao', '竣工日期': 'date',
    'sj_salestaff_name': 'name', 'jzbh_value': 'value', 'gt_zsl': 'gaotao',
    'sj_subs_stat_date': 'date',
})
col_name3 = cols3.get('name', 10)
col_value3 = cols3.get('value', 16)
col_gaotao3 = cols3.get('gaotao', 27)
col_date3 = cols3.get('date', 11)
_exist_has_technical_row = any(
    str(ws3.cell(2, c).value or '').strip() == 'sj_salestaff_name'
    for c in range(1, ws3.max_column + 1)
)
if not all([col_name3, col_value3, col_gaotao3, col_date3]):
    print(f"  存量缺少列: 揽装人={col_name3} 提值={col_value3} 高套={col_gaotao3} 日期={col_date3}")
print(f"  存量列号: 揽装人={col_name3} 提值幅度={col_value3} 高套折算={col_gaotao3} 竣工日期={col_date3}")
_current_dates_exist = []
for r in range(3 if _exist_has_technical_row else 2, ws3.max_row + 1):
    name = str(ws3.cell(r, col_name3).value or '').strip()
    if not name: continue
    tv = safe_float(ws3.cell(r, col_value3).value) or 0
    gt = safe_float(ws3.cell(r, col_gaotao3).value) or 0
    exist_install[name]['value_score'] += tv
    exist_install[name]['gaotao'] += gt
    dv = ws3.cell(r, col_date3).value
    if isinstance(dv, (datetime, date)):
        _current_dates_exist.append(dv if isinstance(dv, date) else dv.date())
wb3.close()

print("4. 读取上月清单...")
last_new = defaultdict(float)
fp10 = os.path.join(DATA_DIR, "上月新装高套清单.xlsx")
wb10 = openpyxl.load_workbook(fp10, data_only=True)
ws10 = wb10.active
_last_dates_new = []
for r in range(3, ws10.max_row + 1):
    name = str(ws10.cell(r, 11).value or '').strip()
    if not name: continue
    pv = safe_float(ws10.cell(r, 15).value) or 0
    last_new[name] += pv
    # 收集竣工日期（列8）
    dv = ws10.cell(r, 8).value
    if isinstance(dv, (datetime, date)):
        _last_dates_new.append(dv if isinstance(dv, date) else dv.date())
wb10.close()

last_exist = defaultdict(float)
fp11 = os.path.join(DATA_DIR, "上月存量高套清单.xlsx")
wb11 = openpyxl.load_workbook(fp11, data_only=True)
ws11 = wb11.active
# 动态检测列号
cols11 = _header_cols(ws11, {
    '揽装人': 'name', '提值幅度': 'value', '竣工日期': 'date',
})
col_name11 = cols11.get('name', 14)
col_value11 = cols11.get('value', 21)
col_date11 = cols11.get('date', 15)
assert col_name11 and col_value11 and col_date11, f"上月存量文件缺少关键列: 揽装人={col_name11} 提值幅度={col_value11} 竣工日期={col_date11}"
print(f"  上月存量列号: 揽装人={col_name11} 提值幅度={col_value11} 竣工日期={col_date11}")
_last_dates_exist = []
for r in range(2, ws11.max_row + 1):
    name = str(ws11.cell(r, col_name11).value or '').strip()
    if not name: continue
    tv = safe_float(ws11.cell(r, col_value11).value) or 0
    last_exist[name] += tv
    dv = ws11.cell(r, col_date11).value
    if isinstance(dv, (datetime, date)):
        _last_dates_exist.append(dv if isinstance(dv, date) else dv.date())
wb11.close()

# ============================================================
# 5. 杠保清单
# ============================================================
print("5. 读取杠保清单...")
gangbao = defaultdict(lambda: {'total': 0, 'success': 0})
fp = os.path.join(DATA_DIR, "杠保清单.xlsx")
wb4 = openpyxl.load_workbook(fp, data_only=True)
ws4 = wb4.active
for r in range(3, ws4.max_row + 1):
    name = str(ws4.cell(r, 8).value or '').strip()
    if not name or name == 'sales_name': continue
    is_gb = safe_float(ws4.cell(r, 20).value) or 0
    gangbao[name]['total'] += 1
    if is_gb == 1:
        gangbao[name]['success'] += 1
wb4.close()

# ============================================================
# 6. 关键一单清单 → 仅装维经理
# ============================================================
print("6. 读取关键一单清单...")
key_order = defaultdict(lambda: {'dispatch': 0, 'convert': 0, 'zhi_dispatch': 0, 'zhi_convert': 0})
fp = os.path.join(DATA_DIR, "关键一单清单.xlsx")
wb5 = openpyxl.load_workbook(fp, data_only=True)
ws5 = wb5.active
cols5 = _header_cols(ws5, {
    '施工人姓名': 'name', '施工人工号': 'name_code',
    '订单状态名称': 'state',
    '实时受理积分': 'real_score',
    '竣工积分': 'finish_score',
    '质差标签': 'zhi_label',
    '综合质差消除校验结果': 'zhi_label2',
})
col_name5 = cols5.get('name') or cols5.get('name_code', 26)
col_state5 = cols5.get('state', 3)
col_real5 = cols5.get('real_score', 101)
col_finish5 = cols5.get('finish_score', 110)
col_zhi5 = cols5.get('zhi_label') or cols5.get('zhi_label2', 129)
for r in range(2, ws5.max_row + 1):
    name = str(ws5.cell(r, col_name5).value or '').strip()
    state = str(ws5.cell(r, col_state5).value or '').strip()
    if not name: continue
    
    # 剔除作废订单
    if state == '作废': continue
    
    key_order[name]['dispatch'] += 1
    
    # 转化判定：实时受理积分>0 OR 竣工积分>0
    real_score = safe_float(ws5.cell(r, col_real5).value) or 0
    finish_score = safe_float(ws5.cell(r, col_finish5).value) or 0
    if real_score > 0 or finish_score > 0:
        key_order[name]['convert'] += 1
    
    # 质差单判定：有质差标签
    zhi_label = str(ws5.cell(r, col_zhi5).value or '').strip()
    if zhi_label:
        key_order[name]['zhi_dispatch'] += 1
        if real_score > 0 or finish_score > 0:
            key_order[name]['zhi_convert'] += 1
wb5.close()

# ============================================================
# 7. 质态相关清单
# ============================================================
print("7. 读取质态相关清单...")
zhitai = defaultdict(lambda: {'t0_invalid': 0, 't0_notrust': 0, 't1_invalid': 0, 't1_notrust': 0, 't3_invalid': 0, 't6_invalid': 0, 'tm1_in': 0, 't2_invalid': 0})
fp = os.path.join(DATA_DIR, "质态相关清单.xlsx")
wb6 = openpyxl.load_workbook(fp, data_only=True)
for sheet_name in wb6.sheetnames:
    if not ('融合质态T+0' in sheet_name or '融合质态T+1' in sheet_name):
        continue
    ws6 = wb6[sheet_name]
    for r in range(3, ws6.max_row + 1):
        name = str(ws6.cell(r, 10).value or '').strip()
        month = str(ws6.cell(r, 3).value or '').strip()
        valid = safe_float(ws6.cell(r, 5).value) or 0
        trust = safe_float(ws6.cell(r, 7).value) or 0
        if not name: continue

        zhitai[name]['tm1_in'] += 1  # T-1 入网

        if month == '202608':
            if valid == 0: zhitai[name]['t0_invalid'] += 1
            if trust == 0: zhitai[name]['t0_notrust'] += 1
        elif month == '202607':
            if valid == 0: zhitai[name]['t1_invalid'] += 1
            if trust == 0: zhitai[name]['t1_notrust'] += 1
        elif month == '202606':
            if valid == 0: zhitai[name]['t2_invalid'] += 1
        elif month == '202605':
            if valid == 0: zhitai[name]['t3_invalid'] += 1
        elif month in ('202602', '202601', '202512'):
            if valid == 0: zhitai[name]['t6_invalid'] += 1
wb6.close()

# ============================================================
# 8. 融合T+0未满卡（质态相关清单 Sheet4）
# ============================================================
print("8. 读取融合T+0未满卡...")
t0_notfull = defaultdict(int)
fp_zt = os.path.join(DATA_DIR, "质态相关清单.xlsx")
wb_zt = openpyxl.load_workbook(fp_zt, data_only=True)
zt_sheets = wb_zt.sheetnames
ws14_idx = [i for i, s in enumerate(zt_sheets) if '未满卡' in s]
if ws14_idx:
    ws14 = wb_zt[zt_sheets[ws14_idx[0]]]
    for r in range(3, ws14.max_row + 1):
        name = str(ws14.cell(r, 8).value or '').strip()
        if not name or name == 'sales_name': continue
        is_full = str(ws14.cell(r, 4).value or '').strip()
        if is_full not in ('是', '1'):
            t0_notfull[name] += 1
wb_zt.close()

#print("All data loaded, generating dashboard...")

# ============================================================
# 9. 聚合数据
# ============================================================
for p in personnel:
    n = p['name']
    ni = new_install.get(n, {})
    ei = exist_install.get(n, {})
    gb = gangbao.get(n, {})
    ko = key_order.get(n, {})
    zt = zhitai.get(n, {})
    
    # 价值积分
    p['new_score'] = round(ni.get('value_score', 0), 2)
    p['exist_score'] = round(ei.get('value_score', 0), 2)
    p['total_score'] = round(p['new_score'] + p['exist_score'], 2)
    p['last_total'] = round(last_new.get(n, 0) + last_exist.get(n, 0), 2)
    
    # 环比（按日均计算）
    # 计算本月/上月数据天数
    _all_cur = _current_dates_new + _current_dates_exist
    _cur_min = min(_all_cur) if _all_cur else None
    _cur_max = max(_all_cur) if _all_cur else None
    _cur_days = (_cur_max - _cur_min).days + 1 if _cur_min else 1
    _all_last = _last_dates_new + _last_dates_exist
    _last_min = min(_all_last) if _all_last else None
    _last_max = max(_all_last) if _all_last else None
    _last_days = 30
    if p['last_total'] and p['last_total'] != 0:
        # 日均环比 = (本月日均 - 上月日均) / |上月日均|
        _cur_daily = p['total_score'] / _cur_days
        _last_daily = p['last_total'] / _last_days
        p['mom'] = round((_cur_daily - _last_daily) / abs(_last_daily), 4)
    else:
        p['mom'] = None
    
    # 高套
    p['new_gaotao'] = round(ni.get('gaotao', 0), 2)
    p['exist_gaotao'] = round(ei.get('gaotao', 0), 2)
    p['total_gaotao'] = round(p['new_gaotao'] + p['exist_gaotao'], 2)
    
    # 杠保
    gb_total = gb.get('total', 0)
    gb_success = gb.get('success', 0)
    p['gb_total'] = gb_total
    p['gb_success'] = gb_success
    p['gb_rate'] = round(gb_success / gb_total, 4) if gb_total > 0 else None
    
    # 关键一单（仅装维经理）
    p['dispatch'] = ko.get('dispatch', 0)
    p['convert'] = ko.get('convert', 0)
    p['convert_rate'] = round(p['convert'] / p['dispatch'], 4) if p['dispatch'] > 0 else None
    p['zhi_dispatch'] = ko.get('zhi_dispatch', 0)
    p['zhi_convert'] = ko.get('zhi_convert', 0)
    if p['role'] != '装维经理':
        p['dispatch'] = None
        p['convert'] = None
        p['convert_rate'] = None
        p['zhi_dispatch'] = None
        p['zhi_convert'] = None
    
    # T+N质态
    p['t1_invalid'] = zt.get('t1_invalid', 0)
    p['t3_invalid'] = zt.get('t3_invalid', 0)
    p['t6_invalid'] = zt.get('t6_invalid', 0)
    
    # 质态相关
    p['t0_notfull'] = t0_notfull.get(n, 0)
    p['t0_invalid'] = zt.get('t0_invalid', 0)
    p['t0_notrust'] = zt.get('t0_notrust', 0)
    p['t1_invalid_2'] = zt.get('t1_invalid', 0)
    p['t1_notrust'] = zt.get('t1_notrust', 0)
    p['tm1_in'] = zt.get('tm1_in', 0)
    
    # T+1有效率
    total_t1 = last_new.get(n, 0) + last_exist.get(n, 0)
    total_t1_count = 1  # placeholder
    # Actually for T+1有效率, need total accounts for this person
    # Let me approximate
    t1_total = zt.get('t1_invalid', 0) + 1  # avoid div by 0
    p['t1_rate'] = round(1 - (zt.get('t1_invalid', 0) / t1_total), 4) if t1_total > 1 else None
    
    # 完成率、缺口
    p['completion_rate'] = round(p['total_gaotao'] / p['target_total'], 4) if p['target_total'] > 0 else None
    p['daliang_gap'] = round(p['target_daliang'] - p['total_gaotao'], 2) if p['target_daliang'] > 0 else None

# ============================================================
# 10. 生成 HTML 看板
# ============================================================
# 10.5. 扫描各数据文件最新日期
# ============================================================
from datetime import date as _dt
def _scan(fp, col, skip=2, sheet_kw=None):
    '''快速扫描文件的最新日期'''
    if not os.path.exists(fp): return '-'
    wb = openpyxl.load_workbook(fp, data_only=True)
    if sheet_kw:
        matches = [s for s in wb.sheetnames if sheet_kw in s]
        ws = wb[matches[0]] if matches else wb.active
    else:
        ws = wb.active
    latest = None
    for r in range(skip, ws.max_row + 1):
        v = ws.cell(r, col).value
        d = None
        if isinstance(v, (_dt, datetime)):
            d = v if isinstance(v, _dt) else v.date()
        elif isinstance(v, int) and 19000000 <= v <= 21000000:
            try: d = datetime.strptime(str(v), '%Y%m%d').date()
            except: pass
        elif isinstance(v, str) and len(v) >= 10:
            try: d = datetime.strptime(v[:19] if ' ' in v[:19] else v[:10], '%Y-%m-%d' if ' ' not in v[:19] else '%Y-%m-%d %H:%M:%S').date()
            except: pass
        if d:
            try: d = d.date()
            except AttributeError: pass
            if d >= _dt(2026, 1, 1) and isinstance(d, _dt):
                if not latest or d > latest: latest = d
    wb.close()
    return latest.strftime('%m/%d') if latest else '-'

_dates = {}
_dates['new_install'] = _scan(os.path.join(DATA_DIR, '新装高套竣工清单.xlsx'), 8, 3)
_dates['exist_install'] = _scan(os.path.join(DATA_DIR, '存量高套竣工清单.xlsx'), 11, 3)
_dates['last_new'] = _scan(os.path.join(DATA_DIR, '上月新装高套清单.xlsx'), 8, 3)
_dates['last_exist'] = _scan(os.path.join(DATA_DIR, '上月存量高套清单.xlsx'), 15, 2)
_dates['gb'] = _scan(os.path.join(DATA_DIR, '杠保清单.xlsx'), 4, 3)
_dates['zt'] = _scan(os.path.join(DATA_DIR, '质态相关清单.xlsx'), 2, 3, sheet_kw='融合质态T+0')
_dates['ko'] = _scan(os.path.join(DATA_DIR, '关键一单清单.xlsx'), 2, 2)
_date_status = f"新装~{_dates['new_install']} 存量~{_dates['exist_install']} 杠保~{_dates['gb']} 质态{_dates['zt']} 关键一单~{_dates['ko']}"
if _shangke_meta:
    _date_status += f" 商客战报~{_shangke_meta.get('report_date', '-')} 数据更新~{_shangke_meta.get('updated_at', '-')}"
else:
    _date_status += " 商客战报~暂无"
# ============================================================
print("9. 生成 HTML 看板...")

def fmt(v, decimals=2):
    """Format a number or display None as -"""
    if v is None: return '-'
    if isinstance(v, float):
        return f"{v:.{decimals}f}"
    return str(v)

def pct(v):
    if v is None: return '-'
    return f"{v*100:.1f}%"

def fmt_score(v):
    if v is None: return '<td class="tr na">-</td>'
    cls = 'p' if v > 0 else ('n' if v < 0 else '')
    return f'<td class="tr {cls}">{v:>8.2f}</td>'

def fmt_mom(v):
    if v is None: return '<td class="tr na">-</td>'
    cls = 'p' if v >= 0 else 'n'
    return f'<td class="tr {cls}">{v*100:>7.1f}%</td>'

html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>沙太人员工作看板 - {data_date[:7]}</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+SC:wght@400;500;600;700&display=swap');
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Noto Sans SC',sans-serif;background:#f0f2f5;color:#1a1a2e;padding:20px}}
.page{{max-width:1400px;margin:0 auto}}
.topbar{{background:linear-gradient(135deg,#0d1b4a,#1a237e,#283593);color:#fff;padding:14px 24px;border-radius:10px;margin-bottom:16px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px}}
.topbar h1{{font-size:18px;letter-spacing:1px;font-weight:700}}
.topbar .meta{{font-size:12px;opacity:.7;display:flex;gap:10px}}
.summary-row{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:16px}}
.summary-card{{background:#fff;border-radius:10px;padding:12px 16px;box-shadow:0 1px 6px rgba(0,0,0,.04);text-align:center}}
.summary-card .num{{font-size:24px;font-weight:700;font-family:'Menlo',monospace}}
.summary-card .label{{font-size:11px;color:#888;margin-top:2px}}

.role-group{{margin-bottom:20px}}
.role-header{{background:#e8eaf6;border-radius:8px;padding:8px 14px;margin-bottom:8px;font-size:14px;font-weight:700;color:#1a237e;display:flex;justify-content:space-between;align-items:center}}
.role-header .rc{{font-size:12px;font-weight:400;color:#666}}

.panel{{background:#fff;border-radius:10px;box-shadow:0 1px 6px rgba(0,0,0,.04);overflow:hidden;margin-bottom:14px}}
.tbl-wrap{{overflow-x:auto}}
table{{width:100%;border-collapse:collapse;font-size:11.5px;min-width:1500px}}
th{{background:#e8eaf6;color:#1a237e;font-weight:600;padding:6px 6px;text-align:center;font-size:11px;position:sticky;top:0;z-index:1;border-right:1px solid #ddd;border-bottom:2px solid #9fa8da;white-space:nowrap}}
th:last-child{{border-right:none}}
td{{padding:4px 4px;border-bottom:1px solid #f0f0f0;text-align:center;white-space:nowrap}}
tr:hover td{{background:#f8f9ff}}
.tr{{text-align:right;font-variant-numeric:tabular-nums;font-family:'Menlo',monospace;font-size:11px}}
.tl{{text-align:left}}
.p{{color:#2e7d32}}
.n{{color:#c62828}}
.na{{color:#ccc}}
.hl-name{{font-weight:500;color:#1a237e}}
.hl-role{{font-size:10px;color:#888;display:block}}
.group-divider td{{background:#f5f7ff!important;font-weight:600;font-size:11px;color:#1a237e;padding:6px 4px}}
.section-title{{font-size:13px;font-weight:700;color:#1a237e;padding:10px 14px 4px;display:flex;align-items:center;gap:6px}}
.cat-header{{background:#283593!important;color:#fff!important;font-size:11px;letter-spacing:1px}}
.cat-sub{{background:#e8eaf6!important;font-size:10px;color:#1a237e;font-weight:600;border-bottom:2px solid #9fa8da!important}}
.rate-bg{{display:inline-block;padding:2px 6px;border-radius:4px;font-weight:600}}
.rate-high{{background:#e8f5e9;color:#2e7d32}}
.rate-mid{{background:#fff3e0;color:#e65100}}
.rate-low{{background:#fce4ec;color:#c62828}}
/* Freeze first 2 columns on horizontal scroll */
td:first-child {{position: sticky; left: 0; z-index: 3; background: #fff; border-right: 1px solid #e0e0e0!important;}}
th:first-child {{position: sticky; left: 0; z-index: 3; border-right: 1px solid #e0e0e0!important;}}
td:nth-child(2) {{position: sticky; left: 82px; z-index: 2; background: #fff; border-right: 1px solid #e0e0e0!important;}}
th:nth-child(2) {{position: sticky; left: 82px; z-index: 2; border-right: 1px solid #e0e0e0!important;}}
thead th:first-child, thead th:nth-child(2) {{ z-index: 3; }}
.cat-header th:first-child, .cat-header th:nth-child(2) {{background: #283593 !important;}}
.footer{{text-align:center;font-size:11px;color:#bbb;margin:16px 0 4px}}

.main-panel{{display:none}}.main-panel.active{{display:block}}
.main-tab-btn{{padding:8px 20px;border:none;background:transparent;border-radius:6px;font-size:13px;font-weight:600;cursor:pointer;color:#666;font-family:inherit;white-space:nowrap}}
.main-tab-btn:hover{{background:rgba(255,255,255,.7);color:#1a237e;transform:translateY(-1px);transition:all .15s;box-shadow:0 1px 3px rgba(0,0,0,.06)}}
.main-tab-btn.active{{background:#fff;color:#1a237e;box-shadow:0 1px 4px rgba(0,0,0,.08)}}

/* UI 优化 */
tr:nth-child(even) td:nth-child(n+3){{background:#fafbff}}
.summary-card:hover{{box-shadow:0 3px 14px rgba(0,0,0,.08);transform:translateY(-1px);transition:all .2s}}
.panel:hover{{box-shadow:0 2px 10px rgba(0,0,0,.06);transition:all .2s}}
</style>
<script>
function switchMainTab(i){{
  document.querySelectorAll(".main-panel").forEach(function(p){{p.classList.remove("active")}});
  document.querySelectorAll(".main-tab-btn").forEach(function(b){{b.classList.remove("active")}});
  var panels = document.querySelectorAll(".main-panel");
  if(panels[i]) panels[i].classList.add("active");
  var btns = document.querySelectorAll(".main-tab-btn");
  if(btns[i]) btns[i].classList.add("active");
}}
</script>
</head>
<body>
<div class="page">
<div class="topbar">
  <h1>📊 沙太人员工作看板</h1>
  <div class="meta"><span>📅 {data_date[:7]}</span><span>👤 {len(personnel)} 人</span></div>
  <div class="date-status" style="width:100%;font-size:10px;color:rgba(255,255,255,.7);margin-top:5px;display:flex;gap:8px;flex-wrap:wrap;border-top:1px solid rgba(255,255,255,.1);padding-top:4px;">
    📡 {_date_status}
  </div>
</div>

<div class="main-tabs" style="display:flex;gap:2px;background:#e8eaf6;border-radius:8px;padding:3px;margin-bottom:14px;"><button class="main-tab-btn active" onclick="switchMainTab(0)">📊 数据对标</button><button class="main-tab-btn" onclick="switchMainTab(1)">📋 底层数据</button></div><div class="main-panel active" id="panel-b"><div class="summary-row">
  <div class="summary-card"><div class="num" style="color:#1565c0;">{len(personnel)}</div><div class="label">总人数</div></div>
  <div class="summary-card"><div class="num" style="color:#2e7d32;">{sum(1 for p in personnel if p['role']=='装维经理')}</div><div class="label">装维经理</div></div>
  <div class="summary-card"><div class="num" style="color:#e65100;">{sum(1 for p in personnel if p['role']=='片区经理')}</div><div class="label">片区经理</div></div>
  <div class="summary-card"><div class="num" style="color:#7b1fa2;">{sum(1 for p in personnel if p['role']=='营业员')}</div><div class="label">营业员</div></div>
</div>
'''

# Build table for each role group
for role_name in ['装维经理', '片区经理', '营业员']:
    group = [p for p in personnel if p['role'] == role_name]
    if not group:
        continue
    
    html += f'<div class="panel"><div class="role-header">{role_name}<span class="rc">{len(group)}人</span></div><div class="tbl-wrap"><table>'
    
    # Table headers - 3 rows: cat header, sub header, column names
    sk_header_html = ''
    sk_sub_html = ''
    if role_name in ('装维经理', '片区经理'):
        sk_header_html = '\n  <th colspan="3">⑦ 商客发展情况</th>'
        sk_sub_html = '\n  <th>增存商企</th><th>小微ICT</th><th>299+发展量</th>'
    html += f'''<thead>
<tr class="cat-header">
  <th rowspan="2" style="min-width:80px;">姓名</th>
  <th rowspan="2" style="min-width:50px;">岗位</th>
  <th colspan="5">① 价值积分</th>
  <th colspan="7">② 增存高套</th>
  <th colspan="3">③ 杠保</th>
  <th colspan="3">④ 关键一单</th>
  <th colspan="5">⑥ 质态相关</th>
{sk_header_html}
</tr>
<tr class="cat-sub">
  <th>新装积分</th><th>存量积分</th><th>合计</th><th>上月</th><th>环比</th>
  <th>新装折算</th><th>存量折算</th><th>总计</th><th>目标</th><th>完成率</th><th>达量目标</th><th>缺口</th>
  <th>杠保成功量</th><th>总业务量</th><th>成功率</th>
  <th>派单</th><th>转化</th><th>转化率</th><th>T0未满卡</th><th>T0无效</th><th>T0无托收</th><th>T1无效</th><th>T1无托收</th>
{sk_sub_html}
</tr></thead><tbody>'''
    
    for p in group:
        rate_cls = ''
        if p['completion_rate'] is not None:
            rate_cls = 'rate-high' if p['completion_rate'] >= 0.8 else ('rate-mid' if p['completion_rate'] >= 0.5 else 'rate-low')
        
        html += f'<tr>'
        html += f'<td class="tl hl-name">{p["name"]}</td>'
        html += f'<td class="hl-role">{p["role"]}</td>'
        
        # ① 价值积分
        html += fmt_score(p['new_score'])
        html += fmt_score(p['exist_score'])
        html += fmt_score(p['total_score'])
        html += fmt_score(p['last_total'])
        html += fmt_mom(p['mom'])
        
        # ② 增存高套
        html += fmt_score(p['new_gaotao'])
        html += fmt_score(p['exist_gaotao'])
        html += fmt_score(p['total_gaotao'])
        html += f'<td class="tr">{p["target_total"]:.1f}</td>'
        if p['completion_rate'] is not None:
            html += f'<td class="tr"><span class="rate-bg {rate_cls}">{pct(p["completion_rate"])}</span></td>'
        else:
            html += '<td class="tr na">-</td>'
        html += f'<td class="tr">{p["target_daliang"]:.1f}</td>'
        html += fmt_score(p['daliang_gap'])
        
        # ③ 杠保
        html += f'<td class="tr">{p["gb_success"]}</td>'
        html += f'<td class="tr">{p["gb_total"]}</td>'
        if p["gb_rate"] is not None:
            r = p["gb_rate"]
            cls = "rate-high" if r >= 0.5 else "rate-mid" if r >= 0.2 else "rate-low"
            html += f'<td class="tr"><span class="rate-bg {cls}">{r*100:.1f}%</span></td>'
        else:
            html += '<td class="tr na">-</td>'
        
        # ④ 关键一单
        if p['dispatch'] is not None:
            html += f'<td class="tr">{p["dispatch"]}</td>'
            html += f'<td class="tr">{p["convert"]}</td>'
            html += f'<td class="tr">{pct(p["convert_rate"])}</td>'
        else:
            html += '<td class="tr na">-</td>' * 3
        
        # ⑥ 质态相关
        zt_s = ' style="background:#ffecb3;color:#e65100;font-weight:700"'
        html += f'<td class="tr"{zt_s if p["t0_notfull"] > 0 else ""}>{p["t0_notfull"]}</td>'
        html += f'<td class="tr"{zt_s if p["t0_invalid"] > 0 else ""}>{p["t0_invalid"]}</td>'
        html += f'<td class="tr"{zt_s if p["t0_notrust"] > 0 else ""}>{p["t0_notrust"]}</td>'
        html += f'<td class="tr"{zt_s if p["t1_invalid_2"] > 0 else ""}>{p["t1_invalid_2"]}</td>'
        html += f'<td class="tr"{zt_s if p["t1_notrust"] > 0 else ""}>{p["t1_notrust"]}</td>'
        
        # ⑦ 商客发展情况（营业员无战报记录，不展示）
        if role_name in ('装维经理', '片区经理'):
            if p['sk_shangqi'] is None:
                html += '<td class="tr na">-</td>' * 3
            else:
                html += f'<td class="tr">{int(p["sk_shangqi"])}</td>'
                html += fmt_score(p['sk_weixiao'])
                html += f'<td class="tr">{int(p["sk_dev299"])}</td>'

        html += '</tr>'
    
    html += '</tbody></table></div></div>'

    # === 数据分析总结（纵向排列，按指标分组）===
    # 辅助函数：给排序列表前2名和后2名
    def _top2(arr, key_fn, reverse=True):
        s = sorted([p for p in arr if key_fn(p) is not None], key=key_fn, reverse=reverse)
        return s[:2], list(reversed(s[-2:])) if len(s) > 2 else (s[:2], [])
    def _fmt(p, key_fn, suffix=''):
        v = key_fn(p)
        if isinstance(v, float) and v == int(v): v = int(v)
        return f'<b>{p["name"]}</b> {v}{suffix}'
    rows = []
    
    # 1. 价值积分
    best_s = max(group, key=lambda p: p['total_score'])
    worst_s = min(group, key=lambda p: p['total_score'])
    rows.append('<div style="margin-bottom:8px;">')
    rows.append('<div style="font-weight:600;color:#1a237e;border-bottom:1px solid #e8eaf6;padding:2px 0;margin-bottom:3px;"><span style="display:inline-block;width:18px;height:18px;background:#1a237e;color:#fff;border-radius:3px;text-align:center;line-height:18px;font-size:10px;margin-right:5px;">' + chr(9312) + '</span>价值积分</div>')
    top2_s, bot2_s = _top2(group, lambda p: p['total_score'])
    if top2_s:
        rows.append('<div style="display:flex;gap:16px;padding-left:24px;"><span style="color:#2e7d32;">▲ 领先:</span>')
        rows.append(' &nbsp; '.join([_fmt(p, lambda x:x['total_score'], '分') for p in top2_s]))
        rows.append('</div>')
    if bot2_s:
        rows.append('<div style="display:flex;gap:16px;padding-left:24px;"><span style="color:#c62828;">▼ 落后:</span>')
        rows.append(' &nbsp; '.join([_fmt(p, lambda x:x['total_score'], '分') for p in bot2_s]))
        rows.append('</div>')
    rows.append('</div>')
    
    # 2. 增存高套
    rows.append('<div style="margin-bottom:8px;">')
    rows.append('<div style="font-weight:600;color:#1a237e;border-bottom:1px solid #e8eaf6;padding:2px 0;margin-bottom:3px;"><span style="display:inline-block;width:18px;height:18px;background:#1a237e;color:#fff;border-radius:3px;text-align:center;line-height:18px;font-size:10px;margin-right:5px;">' + chr(9313) + '</span>增存高套</div>')
    top2_gt, bot2_gt = _top2(group, lambda p: p['total_gaotao'])
    if top2_gt:
        rows.append('<div style="display:flex;gap:16px;padding-left:24px;"><span style="color:#2e7d32;">▲ 领先:</span>')
        rows.append(' &nbsp; '.join([_fmt(p, lambda x:x['total_gaotao']) for p in top2_gt]))
        rows.append('</div>')
    if bot2_gt:
        rows.append('<div style="display:flex;gap:16px;padding-left:24px;"><span style="color:#c62828;">▼ 落后:</span>')
        rows.append(' &nbsp; '.join([_fmt(p, lambda x:x['total_gaotao']) for p in bot2_gt]))
        rows.append('</div>')
    rows.append('</div>')
    
    # 3. 杠保
    gb_valid = [p for p in group if p['gb_rate'] is not None]
    if gb_valid:
        rows.append('<div style="margin-bottom:8px;">')
        rows.append('<div style="font-weight:600;color:#1a237e;border-bottom:1px solid #e8eaf6;padding:2px 0;margin-bottom:3px;"><span style="display:inline-block;width:18px;height:18px;background:#1a237e;color:#fff;border-radius:3px;text-align:center;line-height:18px;font-size:10px;margin-right:5px;">' + chr(9314) + '</span>杠保成功率</div>')
        top2_gb = sorted(gb_valid, key=lambda p: p['gb_rate'], reverse=True)[:2] if gb_valid else []
        bot2_gb = sorted(gb_valid, key=lambda p: p['gb_rate'])[:2] if gb_valid else []
        if top2_gb:
            rows.append('<div style="display:flex;gap:16px;padding-left:24px;"><span style="color:#2e7d32;">▲ 领先:</span>')
            rows.append(' &nbsp; '.join([_fmt(p, lambda x:round(x['gb_rate']*100,1), '%') for p in top2_gb]))
            rows.append('</div>')
        if bot2_gb:
            rows.append('<div style="display:flex;gap:16px;padding-left:24px;"><span style="color:#c62828;">▼ 落后:</span>')
            rows.append(' &nbsp; '.join([_fmt(p, lambda x:round(x['gb_rate']*100,1), '%') for p in bot2_gb]))
            rows.append('</div>')
        rows.append('</div>')
    
    # 4. 关键一单（仅装维经理）
    if role_name == '装维经理':
        ko_valid = [p for p in group if p['convert_rate'] is not None]
        if ko_valid:
            rows.append('<div style="margin-bottom:8px;">')
            rows.append('<div style="font-weight:600;color:#1a237e;border-bottom:1px solid #e8eaf6;padding:2px 0;margin-bottom:3px;"><span style="display:inline-block;width:18px;height:18px;background:#1a237e;color:#fff;border-radius:3px;text-align:center;line-height:18px;font-size:10px;margin-right:5px;">' + chr(9315) + '</span>关键一单转化率<span style="font-size:10px;color:#999;font-weight:400;margin-left:4px;">（仅装维经理）</span></div>')
            top2_ko, bot2_ko = _top2(ko_valid, lambda p: p['convert_rate'])
            if top2_ko:
                rows.append('<div style="display:flex;gap:16px;padding-left:24px;"><span style="color:#2e7d32;">▲ 领先:</span>')
                rows.append(' &nbsp; '.join([_fmt(p, lambda x:round(x['convert_rate']*100,1), '%') + f'（派单{p["dispatch"]}、转化{p["convert"]}）' for p in top2_ko]))
                rows.append('</div>')
            if bot2_ko:
                rows.append('<div style="display:flex;gap:16px;padding-left:24px;"><span style="color:#c62828;">▼ 落后:</span>')
                rows.append(' &nbsp; '.join([_fmt(p, lambda x:round(x['convert_rate']*100,1), '%') + f'（派单{p["dispatch"]}、转化{p["convert"]}）' for p in bot2_ko]))
                rows.append('</div>')
            rows.append('</div>')
    
    # 5. 质态关注
    zt_total = lambda p: (p.get('t0_notfull', 0) + p.get('t0_invalid', 0) + p.get('t0_notrust', 0) + p.get('t1_invalid_2', 0) + p.get('t1_notrust', 0))
    zt_people = sorted(group, key=lambda p: zt_total(p), reverse=True)
    zt_need = [p for p in zt_people if zt_total(p) > 0]
    if zt_need:
        rows.append('<div style="margin-bottom:2px;">')
        rows.append('<div style="font-weight:600;color:#1a237e;border-bottom:1px solid #e8eaf6;padding:2px 0;margin-bottom:3px;"><span style="display:inline-block;width:18px;height:18px;background:#1a237e;color:#fff;border-radius:3px;text-align:center;line-height:18px;font-size:10px;margin-right:5px;">' + chr(9316) + '</span>质态关注（需要提醒的人员）</div>')
        for p in zt_need[:2]:
            fields = []
            if p['t0_notfull'] > 0: fields.append(f'T0未满卡{p["t0_notfull"]}')
            if p['t0_invalid'] > 0: fields.append(f'T0无效{p["t0_invalid"]}')
            if p['t0_notrust'] > 0: fields.append(f'T0无托收{p["t0_notrust"]}')
            if p['t1_invalid_2'] > 0: fields.append(f'T1无效{p["t1_invalid_2"]}')
            if p['t1_notrust'] > 0: fields.append(f'T1无托收{p["t1_notrust"]}')
            detail = '、'.join(fields)
            rows.append(f'<div style="display:flex;gap:16px;padding-left:24px;"><span style="color:#c62828;">⚠️ <b>{p["name"]}</b> 共{zt_total(p)}项异常：{detail}</span></div>')
        rows.append('</div>')

    # 商客发展未破零（商企/小微/299+ 均为0）
    if role_name in ('装维经理', '片区经理'):
        sk_zero = [p for p in group if p.get('sk_shangqi') is not None and p.get('sk_shangqi') == 0
                   and p.get('sk_weixiao') == 0 and p.get('sk_dev299') == 0]
        if sk_zero:
            rows.append('<div style="margin-bottom:8px;">')
            rows.append('<div style="font-weight:600;color:#1a237e;border-bottom:1px solid #e8eaf6;padding:2px 0;margin-bottom:3px;"><span style="display:inline-block;width:18px;height:18px;background:#1a237e;color:#fff;border-radius:3px;text-align:center;line-height:18px;font-size:10px;margin-right:5px;">' + chr(9317) + '</span>商客发展未破零<span style="font-size:10px;color:#999;font-weight:400;margin-left:4px;">（商企/小微/299+均为0）</span></div>')
            names = ' &nbsp; '.join([f'<span style="color:#c62828;"><b>{p["name"]}</b></span>' for p in sk_zero])
            rows.append(f'<div style="display:flex;gap:12px;padding-left:24px;flex-wrap:wrap;">{names}</div>')
            rows.append('</div>')
    
    # 生成分析 HTML
    if rows:
        html += '<div class="panel" style="padding:10px 14px;font-size:11px;color:#444;">'
        html += f'<div style="font-weight:700;color:#1a237e;border-bottom:2px solid #e8eaf6;padding-bottom:4px;margin-bottom:6px;">📊 {role_name} 各指标领先落后分析</div>'
        html += ''.join(rows)
        html += '</div>'
    
# Legend
html += '''
<div class="panel" style="padding:10px 14px;font-size:11px;color:#666;">
  <b>说明：</b>
  <span style="color:#2e7d32;">■ 绿色</span> = 正值/达标(≥80%)
  <span style="color:#e65100;margin-left:12px;">■ 橙色</span> = 中等(50-80%)
  <span style="color:#c62828;margin-left:12px;">■ 红色</span> = 负值/低(<50%)
  <span style="color:#ccc;margin-left:12px;">—</span> = 不适用/无数据
  <span style="margin-left:12px;">📌 关键一单仅装维经理对标</span>
</div>
</div> <!-- close panel-b -->
'''

# === 底层数据只读面板 ===
_mgmt_rows = []
for _i, _p in enumerate(personnel, 1):
    _code = str(_p.get('code', '') or '').strip()
    _sk_rec = _shangke.get(_code)
    _sk_status = '✓' if _sk_rec else '-'
    _mgmt_rows.append(
        f'<tr><td>{_i}</td><td class="tl hl-name">{_p["name"]}</td>'
        f'<td>{_code}</td><td>{_p["role"]}</td><td>{_p.get("cp", "") or "-"}</td>'
        f'<td class="tr">{_p.get("target_total", 0):.1f}</td>'
        f'<td class="tr">{_p.get("target_daliang", 0):.1f}</td>'
        f'<td class="tr">{_sk_status}</td></tr>'
    )
html += f'''<div class="main-panel" id="panel-m">
<div class="panel" style="margin-top:10px;">
<div style="display:flex;justify-content:space-between;align-items:center;padding:10px 14px;border-bottom:2px solid #c5cae9;margin-bottom:10px;">
<span style="font-size:14px;font-weight:700;color:#1a237e;">📋 底层数据</span>
<span style="font-size:11px;color:#888;">只读展示人员/编码/岗位/CP/目标，更新请使用月度模板</span>
</div>
<div class="tbl-wrap"><table style="min-width:auto;font-size:12px;">
<thead><tr><th>#</th><th class="tl">姓名</th><th>揽装编码</th><th>岗位</th><th>CP</th><th>增存高套目标</th><th>达量目标</th><th>商客战报</th></tr></thead>
<tbody>{''.join(_mgmt_rows)}</tbody></table></div></div></div>
'''

html += f'''
<div class="footer">沙太人员工作看板 · {data_date[:7]} · 数据源: {int(data_date[5:7])}月清单</div>
</div>
</body>
</html>'''

os.makedirs(OUTPUT_DIR, exist_ok=True)
outpath = os.path.join(OUTPUT_DIR, "dashboard.html")
with open(outpath, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"\n✅ 看板已生成: {outpath}")
print(f"   文件大小: {os.path.getsize(outpath) / 1024:.0f} KB")
