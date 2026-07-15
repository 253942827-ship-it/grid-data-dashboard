import openpyxl, os, json
from collections import defaultdict

PROJ = "/Users/mr.g/Documents/Codex/Workspace/projects/project-006-number-lifecycle-monitor"
DATA = os.path.join(PROJ, "data")
code_map = {'100000':'在用','110009':'预拆机','120000':'停机'}
t_tags = ['T+0','T+1','T+2','T+3','T+4','T+5','T+6','T+7','T+8','T+9','T+10','T+11','T+12','T+13']

# 读跟踪表
wb = openpyxl.load_workbook(os.path.join(DATA, "号码状态跟踪_2025年5月批次.xlsx"), data_only=True)
ws = wb.active
records = []
for r in range(2, ws.max_row + 1):
    rec = {'date':str(ws.cell(r,1).value or '')[:10],'phone':str(ws.cell(r,2).value or ''),
           'bb':str(ws.cell(r,3).value or ''),'person':str(ws.cell(r,5).value or ''),'project':str(ws.cell(r,6).value or '')}
    for i in range(14):
        rec[f's{i}'] = str(ws.cell(r,8+i*2+1).value or '')
    records.append(rec)
wb.close()

# 累计价值积分
vp = defaultdict(float)
for fn in sorted([f for f in os.listdir(DATA) if f.startswith('25年5月清单')]):
    wb2 = openpyxl.load_workbook(os.path.join(DATA, fn), data_only=True)
    ws2 = wb2.active
    for r in range(2, ws2.max_row + 1):
        bb = str(ws2.cell(r,1).value or '').strip()
        if not bb or bb == '<null>': continue
        v = ws2.cell(r,25).value
        if isinstance(v,(int,float)): vp[bb] += round(v,2)
    wb2.close()
for rec in records:
    rec['vp'] = vp.get(rec['bb'],0)

# 项目分组
projects = {}
for p in ['花生寮','鸿辉','合神','冠均']:
    ps = [r for r in records if r['project']==p]
    if ps: projects[p] = ps

def st(recs):
    i = sum(1 for r in recs if r['s13']=='在用')
    s = sum(1 for r in recs if r['s13']=='停机')
    p = sum(1 for r in recs if r['s13']=='预拆机')
    v = round(sum(r['vp'] for r in recs),2)
    return {'t':len(recs),'i':i,'s':s,'p':p,'v':v,'ir':f'{i/len(recs)*100:.0f}%'}

# 构建HTML
H = lambda: None
h = ''
def a(s): global h; h += s + '\n'

a('<!DOCTYPE html>')
a('<html lang="zh-CN">')
a('<head>')
a('<meta charset="UTF-8">')
a('<meta name="viewport" content="width=device-width,initial-scale=1.0">')
a('<title>号码生命周期 - 监控看板</title>')
a('<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>')
a('<style>')
a('*{margin:0;padding:0;box-sizing:border-box}')
a('body{font-family:-apple-system,sans-serif;background:#f0f2f5;color:#1a1a2e;padding:20px}')
a('.page{max-width:1300px;margin:0 auto}')
a('.topbar{background:linear-gradient(135deg,#0d1b4a,#1a237e);color:#fff;padding:14px 24px;border-radius:10px;margin-bottom:16px}')
a('.topbar h1{font-size:18px;font-weight:700}')
a('.topbar .meta{font-size:12px;opacity:.7;margin-top:4px}')
a('.summary{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:10px;margin-bottom:16px}')
a('.card{background:#fff;border-radius:10px;padding:14px;text-align:center;box-shadow:0 1px 4px rgba(0,0,0,.04)}')
a('.card .num{font-size:24px;font-weight:700}')
a('.card .label{font-size:11px;color:#888;margin-top:2px}')
a('.panel{background:#fff;border-radius:10px;padding:14px;box-shadow:0 1px 4px rgba(0,0,0,.04);margin-bottom:14px}')
a('.panel-title{font-size:14px;font-weight:600;color:#1a237e;padding-bottom:8px;border-bottom:2px solid #e8eaf6;margin-bottom:10px}')
a('.tab-bar{display:flex;gap:2px;background:#e8eaf6;border-radius:8px;padding:3px;margin-bottom:14px;flex-wrap:wrap}')
a('.tab-btn{padding:7px 16px;border:none;border-radius:6px;font-size:12px;font-weight:600;cursor:pointer;color:#666;background:transparent;font-family:inherit}')
a('.tab-btn:hover{background:rgba(255,255,255,.7);color:#1a237e}')
a('.tab-btn.active{background:#fff;color:#1a237e;box-shadow:0 1px 3px rgba(0,0,0,.08)}')
a('.tab-panel{display:none}')
a('.tab-panel.active{display:block}')
a('table{width:100%;border-collapse:collapse;font-size:11.5px}')
a('th{background:#e8eaf6;color:#1a237e;padding:5px 6px;text-align:center;font-size:11px;border-bottom:2px solid #9fa8da;white-space:nowrap}')
a('td{padding:4px 5px;text-align:center;border-bottom:1px solid #f0f0f0;white-space:nowrap;font-size:11px}')
a('.tr{text-align:right}')
a('.tag{display:inline-block;padding:1px 8px;border-radius:4px;font-size:10px;font-weight:500}')
a('.green{background:#e8f5e9;color:#2e7d32}')
a('.red{background:#fce4ec;color:#c62828}')
a('.tbl-wrap{overflow-x:auto;max-height:65vh}')
a('.search-bar input{padding:5px 10px;border:1px solid #ddd;border-radius:6px;width:200px;font-size:12px;margin-bottom:8px}')
a('.chart-box{height:260px}')
a('.status-bar{display:flex;height:20px;border-radius:3px;overflow:hidden;margin:3px 0}')
a('.status-bar div{transition:width .5s}')
a('.small{font-size:10px;color:#999}')
a('</style></head><body>')
a('<div class="page">')
a('<div class="topbar"><h1>📊 号码生命周期 · 项目监控看板</h1><div class="meta">2025年5月入网批次 · 共' + str(len(records)) + '条号码</div></div>')

# 概览卡片
all_s = st(records)
a('<div class="summary">')
a(f'<div class="card"><div class="num" style="color:#1565c0">{all_s["t"]}</div><div class="label">监控号码</div></div>')
a(f'<div class="card"><div class="num" style="color:#2e7d32">{all_s["i"]}</div><div class="label">在用(T+13)</div></div>')
a(f'<div class="card"><div class="num" style="color:#e65100">{all_s["s"]}</div><div class="label">停机(T+13)</div></div>')
a(f'<div class="card"><div class="num" style="color:#6a1b9a">{all_s["v"]}</div><div class="label">累计价值积分</div></div>')
a('</div>')

# 趋势图
all_i = [sum(1 for r in records if r[f's{i}']=='在用') for i in range(14)]
all_sv = [sum(1 for r in records if r[f's{i}']=='停机') for i in range(14)]
all_p = [sum(1 for r in records if r[f's{i}']=='预拆机') for i in range(14)]
a('<div class="panel"><div class="panel-title">📈 状态演变趋势</div><div class="chart-box"><canvas id="trendChart"></canvas></div></div>')

# 标签
a('<div class="tab-bar">')
a(f'<button class="tab-btn active" onclick="switchTab(0,\'overview\')">📊 总览</button>')
names = ['overview'] + list(projects.keys())
for i, p in enumerate(names[1:], 1):
    a(f'<button class="tab-btn" onclick="switchTab({i},\'{p}\')">{p}</button>')
a('</div>')

# 总览标签
a('<div class="tab-panel active" id="tab-overview">')
a('<div class="panel"><div class="panel-title">📋 各项目汇总</div>')
a('<div class="tbl-wrap"><table><thead><tr><th>项目</th><th>号码</th><th>在用</th><th>停机</th><th>预拆机</th><th>在用率</th><th>累计积分</th><th>状态</th></tr></thead><tbody>')
for p, recs in projects.items():
    s = st(recs)
    a(f'<tr><td><b>{p}</b></td><td>{s["t"]}</td><td style="color:#2e7d32">{s["i"]}</td><td style="color:#e65100">{s["s"]}</td><td style="color:#c62828">{s["p"]}</td><td>{s["ir"]}</td><td class="tr">{s["v"]}</td>'
      f'<td><div class="status-bar"><div style="width:{s["i"]/s["t"]*100}%;background:#2e7d32"></div>'
      f'<div style="width:{s["s"]/s["t"]*100}%;background:#e65100"></div>'
      f'<div style="width:{s["p"]/s["t"]*100}%;background:#c62828"></div></div></td></tr>')
a('</tbody></table></div></div></div>')

# 各项目标签
for p, recs in projects.items():
    s = st(recs)
    a(f'<div class="tab-panel" id="tab-{p}">')
    a(f'<div class="summary" style="grid-template-columns:repeat(4,1fr)">')
    a(f'<div class="card"><div class="num" style="color:#1565c0">{s["t"]}</div><div class="label">号码</div></div>')
    a(f'<div class="card"><div class="num" style="color:#2e7d32">{s["i"]}</div><div class="label">在用(T+13)</div></div>')
    a(f'<div class="card"><div class="num" style="color:#e65100">{s["s"]}</div><div class="label">停机(T+13)</div></div>')
    a(f'<div class="card"><div class="num" style="color:#6a1b9a">{s["v"]}</div><div class="label">累计积分</div></div>')
    a('</div>')
    a('<div class="panel"><div class="panel-title">📋 号码详情 · T+0~T+13 状态链</div>')
    a(f'<div class="search-bar"><input type="text" id="search_{p}" oninput="filterTable(this,\'{p}\')" placeholder="🔍 搜索号码..."></div>')
    a('<div class="tbl-wrap"><table><thead><tr><th>#</th><th>入网</th><th>号码</th><th>揽装人</th><th>积分</th>')
    for t in t_tags: a(f'<th>{t}</th>')
    a('</tr></thead><tbody>')
    for idx, rec in enumerate(recs):
        a(f'<tr><td>{idx+1}</td><td>{rec["date"]}</td><td><b>{rec["phone"]}</b></td><td>{rec["person"]}</td><td class="tr">{rec["vp"]}</td>')
        for i in range(14):
            s_tag = rec[f's{i}']
            cls = 'green' if s_tag == '在用' else 'red' if s_tag in ('停机','预拆机') else ''
            a(f'<td><span class="tag {cls}" style="font-size:9px;padding:1px 4px">{s_tag[:2] if s_tag else "-"}</span></td>')
        a('</tr>')
    a('</tbody></table></div></div></div>')

a('</div>')
a('<script>')
a(f'new Chart(document.getElementById("trendChart"),{{type:"line",data:{{labels:{json.dumps(t_tags)},datasets:[')
a(f'{{label:"在用",data:{json.dumps(all_i)},borderColor:"#2e7d32",backgroundColor:"rgba(46,125,50,.08)",fill:true,tension:.3,pointRadius:3}},')
a(f'{{label:"停机",data:{json.dumps(all_sv)},borderColor:"#e65100",backgroundColor:"rgba(230,81,0,.08)",fill:true,tension:.3,pointRadius:3}},')
a(f'{{label:"预拆机",data:{json.dumps(all_p)},borderColor:"#c62828",backgroundColor:"rgba(198,40,40,.08)",fill:true,tension:.3,pointRadius:3}}')
a(']}},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:"top"}},scales:{y:{beginAtZero:true}}}});')
a('function switchTab(i,n){document.querySelectorAll(".tab-panel,.tab-btn").forEach(function(e){e.classList.remove("active")});document.getElementById("tab-"+n).classList.add("active");document.querySelectorAll(".tab-btn")[i].classList.add("active")}')
a('function filterTable(inp,p){var q=inp.value;document.querySelectorAll("#tab-"+p+" tbody tr").forEach(function(r){r.style.display=r.textContent.includes(q)?"":"none"})}')
a('</script>')
a('</body></html>')

fp = os.path.join(PROJ, 'docs', 'monitor.html')
os.makedirs(os.path.dirname(fp), exist_ok=True)
with open(fp, 'w', encoding='utf-8') as f:
    f.write(h)
print(f"✅ 看板已生成 ({len(h)/1024:.0f} KB, {len(h.split(chr(10)))} 行)")
