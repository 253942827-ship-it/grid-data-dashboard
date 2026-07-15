#!/usr/bin/env python3
import openpyxl, os, json, re

PROJ = "/Users/mr.g/Documents/Codex/Workspace/projects/project-006-number-lifecycle-monitor"
DATA = os.path.join(PROJ, "data")
OUT = os.path.join(PROJ, "docs", "monitor.html")
code_map = {'100000':'在用','110009':'预拆机','120000':'停机'}

# 1. 读跟踪表
fp = os.path.join(DATA, "号码状态跟踪_2025年5月批次.xlsx")
wb = openpyxl.load_workbook(fp, data_only=True)
ws = wb.active
t_tags = ['T+0','T+1','T+2','T+3','T+4','T+5','T+6','T+7','T+8','T+9','T+10','T+11','T+12','T+13']
records = []
for r in range(2, ws.max_row + 1):
    rec = {'date':str(ws.cell(r,1).value or '')[:10],'phone':str(ws.cell(r,2).value or ''),
           'bb':str(ws.cell(r,3).value or ''),'pkg':str(ws.cell(r,4).value or ''),
           'person':str(ws.cell(r,5).value or ''),'project':str(ws.cell(r,6).value or ''),'tm':str(ws.cell(r,7).value or '')}
    for i in range(14):
        rec[f'b{i}'] = str(ws.cell(r, 8+i*2).value or '')
        rec[f's{i}'] = str(ws.cell(r, 8+i*2+1).value or '')
    records.append(rec)
wb.close()

# 2. 读最新状态文件获取加载时间和价值积分
latest_fp = os.path.join(DATA, "25年5月清单在26年6月状态.xlsx")
wb2 = openpyxl.load_workbook(latest_fp, data_only=True)
ws2 = wb2.active
realtime = {}
for r in range(2, ws2.max_row + 1):
    bb = str(ws2.cell(r, 1).value or '').strip()
    if not bb or bb == '<null>': continue
    load_time = str(ws2.cell(r, 7).value or '')
    value_pts = ws2.cell(r, 25).value
    if isinstance(value_pts, (int, float)): value_pts = round(value_pts, 2)
    else: value_pts = 0
    online = ws2.cell(r, 21).value
    status_code = str(ws2.cell(r, 9).value or '')
    realtime[bb] = {'load_time': load_time, 'value_pts': value_pts, 'online': online, 'status': code_map.get(status_code, status_code)}
wb2.close()

# 3. 为每条记录加上最新状态信息
for rec in records:
    rt = realtime.get(rec['bb'], {})
    rec['load_time'] = str(rt.get('load_time',''))[:19]
    rec['value_pts'] = rt.get('value_pts', 0)
    rec['latest_status'] = rt.get('status', '')
    rec['latest_online'] = rt.get('online', '')

# 4. 项目分组
projects_order = ['花生寮','鸿辉','合神','冠均']
projects = {}
for p in projects_order:
    projects[p] = [rec for rec in records if rec['project'] == p]

# 计算各项目统计
proj_stats = {}
for p, recs in projects.items():
    if not recs: continue
    latest = recs[0]['s13']  # T+13
    s_inuse = sum(1 for r in recs if r['s13'] == '在用')
    s_stop = sum(1 for r in recs if r['s13'] == '停机')
    s_pre = sum(1 for r in recs if r['s13'] == '预拆机')
    total_vp = sum(r.get('value_pts',0) for r in recs)
    proj_stats[p] = {'total': len(recs), 'inuse': s_inuse, 'stop': s_stop, 'pre': s_pre, 'vp': total_vp}

# 生成HTML
html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>号码生命周期 - 项目状态监控</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:-apple-system,sans-serif;background:#f0f2f5;color:#1a1a2e;padding:20px}}
.page{{max-width:1300px;margin:0 auto}}
.topbar{{background:linear-gradient(135deg,#0d1b4a,#1a237e);color:#fff;padding:14px 24px;border-radius:10px;margin-bottom:16px}}
.topbar h1{{font-size:18px;font-weight:700}}
.topbar .meta{{font-size:12px;opacity:.7;margin-top:4px}}
.summary{{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:10px;margin-bottom:16px}}
.card{{background:#fff;border-radius:10px;padding:14px;text-align:center;box-shadow:0 1px 4px rgba(0,0,0,.04)}}
.card .num{{font-size:24px;font-weight:700}}
.card .label{{font-size:11px;color:#888;margin-top:2px}}
.card .sub{{font-size:10px;color:#999;margin-top:1px}}
.panel{{background:#fff;border-radius:10px;padding:14px;box-shadow:0 1px 4px rgba(0,0,0,.04);margin-bottom:14px}}
.panel-title{{font-size:14px;font-weight:600;color:#1a237e;padding-bottom:8px;border-bottom:2px solid #e8eaf6;margin-bottom:10px}}
.tab-bar{{display:flex;gap:2px;background:#e8eaf6;border-radius:8px;padding:3px;margin-bottom:14px;flex-wrap:wrap}}
.tab-btn{{padding:7px 16px;border:none;border-radius:6px;font-size:12px;font-weight:600;cursor:pointer;color:#666;background:transparent;font-family:inherit}}
.tab-btn:hover{{background:rgba(255,255,255,.7);color:#1a237e}}
.tab-btn.active{{background:#fff;color:#1a237e;box-shadow:0 1px 3px rgba(0,0,0,.08)}}
.tab-panel{{display:none}}.tab-panel.active{{display:block}}
table{{width:100%;border-collapse:collapse;font-size:11.5px}}
th{{background:#e8eaf6;color:#1a237e;padding:5px 6px;text-align:center;font-size:11px;border-bottom:2px solid #9fa8da;position:sticky;top:0;z-index:1;white-space:nowrap}}
td{{padding:4px 5px;text-align:center;border-bottom:1px solid #f0f0f0;white-space:nowrap;font-size:11px}}
.tl{{text-align:left}}.tr{{text-align:right}}
.tag{{display:inline-block;padding:1px 8px;border-radius:4px;font-size:10px;font-weight:500}}
.green{{background:#e8f5e9;color:#2e7d32}}
.red{{background:#fce4ec;color:#c62828}}
.yellow{{background:#fff3e0;color:#e65100}}
.blue{{background:#e3f2fd;color:#1565c0}}
.tbl-wrap{{overflow-x:auto;max-height:65vh}}
.search-bar{{margin-bottom:8px;display:flex;gap:8px;flex-wrap:wrap;align-items:center}}
.search-bar input{{padding:5px 10px;border:1px solid #ddd;border-radius:6px;width:200px;font-size:12px}}
.proj-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:10px;margin-bottom:14px}}
.project-card{{background:#fff;border-radius:10px;padding:14px;box-shadow:0 1px 4px rgba(0,0,0,.04);border-left:4px solid #1a237e}}
.project-card .name{{font-size:14px;font-weight:600;color:#1a237e;margin-bottom:6px}}
.project-card .stat{{font-size:11px;color:#666;margin:1px 0}}
.project-card .stat span{{font-weight:600;font-size:13px}}
.status-bar{{display:flex;height:20px;border-radius:3px;overflow:hidden;margin:3px 0}}
.status-bar div{{transition:width .5s}}
.small{{font-size:10px;color:#999}}
.chart-box{{height:260px}}
.t-cycle td{{font-size:10px;padding:3px 4px}}
</style></head><body>
<div class="page">
<div class="topbar"><h1>📊 号码生命周期 · 项目状态监控</h1>
<div class="meta">2025年5月入网批次 · 共{len(records)}条号码 · 数据截至2026年7月</div></div>

<div class="summary">
<div class="card"><div class="num" style="color:#1565c0">{len(records)}</div><div class="label">监控号码总数</div><div class="sub">覆盖{len(projects)}个项目</div></div>
<div class="card"><div class="num" style="color:#2e7d32">{sum(p['inuse'] for p in proj_stats.values())}</div><div class="label">当前在用</div><div class="sub">T+13最新状态</div></div>
<div class="card"><div class="num" style="color:#e65100">{sum(p['stop'] for p in proj_stats.values())}</div><div class="label">当前停机</div><div class="sub">T+13最新状态</div></div>
<div class="card"><div class="num" style="color:#c62828">{sum(p['pre'] for p in proj_stats.values())}</div><div class="label">预拆机</div><div class="sub">累计出现</div></div>
<div class="card"><div class="num" style="color:#6a1b9a">{sum(p['vp'] for p in proj_stats.values())}</div><div class="label">当月价值积分</div><div class="sub">26年6月</div></div>
</div>

<div class="proj-grid">'''
for p in projects_order:
    s = proj_stats.get(p, {'total':0,'inuse':0,'stop':0,'pre':0,'vp':0})
    rate = f'{s["inuse"]/max(s["total"],1)*100:.0f}%' if s['total']>0 else '—'
    html += f'''<div class="project-card" style="border-left-color:{'#2e7d32' if rate=='100%' else '#e65100'}">
<div class="name">{p}</div>
<div class="stat">号码: <span>{s['total']}</span> 在用: <span>{s['inuse']}</span> ({rate})</div>
<div class="stat">停机: <span>{s['stop']}</span> 预拆机: <span>{s['pre']}</span> 价值积分: <span>{s['vp']}</span></div>
</div>'''

# T周期趋势汇总
html += '''</div>
<div class="panel">
<div class="panel-title">📈 状态演变趋势</div>
<div class="chart-box"><canvas id="trendChart"></canvas></div></div>'''

# 标签栏
html += '<div class="tab-bar"><button class="tab-btn active" onclick="switchTab(0,\'overview\')">📊 总览</button>'
for i, p in enumerate(projects_order):
    if projects[p]:
        html += f'<button class="tab-btn" onclick="switchTab({i+1},\'{p}\')">{p}</button>'
html += '</div>'

# 总览标签
html += f'''<div class="tab-panel active" id="tab-overview">
<div class="panel"><div class="panel-title">📋 各项目实时状态（26年6月）</div>
<div class="tbl-wrap"><table><thead><tr><th>项目</th><th>总号码</th><th>在用</th><th>停机</th><th>预拆机</th><th>在用率</th><th>价值积分</th><th>状态分布</th></tr></thead><tbody>'''
for p in projects_order:
    s = proj_stats.get(p, {'total':0,'inuse':0,'stop':0,'pre':0,'vp':0})
    html += f'<tr><td><b>{p}</b></td><td>{s["total"]}</td><td class="green">{s["inuse"]}</td><td class="red">{s["stop"]}</td><td class="red">{s["pre"]}</td>'
    html += f'<td>{s["inuse"]/max(s["total"],1)*100:.0f}%</td><td class="tr">{s["vp"]}</td>'
    html += '<td><div class="status-bar">'
    if s['inuse']>0: html += f'<div style="width:{s["inuse"]/max(s["total"],1)*100}%;background:#2e7d32" title="在用{s["inuse"]}"></div>'
    if s['stop']>0: html += f'<div style="width:{s["stop"]/max(s["total"],1)*100}%;background:#e65100" title="停机{s["stop"]}"></div>'
    if s['pre']>0: html += f'<div style="width:{s["pre"]/max(s["total"],1)*100}%;background:#c62828" title="预拆机{s["pre"]}"></div>'
    html += '</div></td></tr>'
html += '</tbody></table></div></div></div>'

# 各项目标签
for p in projects_order:
    recs = projects.get(p, [])
    if not recs: continue
    act = 'active' if p == projects_order[0] else ''
    html += f'<div class="tab-panel" id="tab-{p}">'
    s = proj_stats.get(p, {'total':0,'inuse':0,'stop':0,'pre':0,'vp':0})
    html += f'''<div class="summary" style="grid-template-columns:repeat(4,1fr)">
<div class="card"><div class="num" style="color:#1565c0">{s["total"]}</div><div class="label">号码总数</div></div>
<div class="card"><div class="num" style="color:#2e7d32">{s["inuse"]}</div><div class="label">在用</div></div>
<div class="card"><div class="num" style="color:#e65100">{s["stop"]}</div><div class="label">停机</div></div>
<div class="card"><div class="num" style="color:#c62828">{s["pre"]}</div><div class="label">预拆机</div></div>
</div>'''
    
    html += '''<div class="panel">
<div class="panel-title">📋 号码详情 & 实时状态</div>
<div class="search-bar"><input type="text" id="search-''' + p + '''" oninput="filterProjTable(this,''' + json.dumps(p) + ''')" placeholder="🔍 搜索号码/揽装人..."></div>
<div class="tbl-wrap">
<table><thead><tr><th>#</th><th>入网日期</th><th>号码</th><th>宽带接入号</th><th>揽装人</th><th>T月</th><th>最新状态</th><th>加载时间</th><th>价值积分</th><th>T+0</th><th>T+1</th><th>T+2</th><th>T+3</th><th>T+4</th><th>T+5</th><th>T+6</th><th>T+7</th><th>T+8</th><th>T+9</th><th>T+10</th><th>T+11</th><th>T+12</th><th>T+13</th></tr></thead><tbody id="tbody-''' + p + '">'
    for idx, rec in enumerate(recs):
        html += f'<tr><td>{idx+1}</td><td>{rec["date"]}</td><td><b>{rec["phone"]}</b></td><td class="small">{rec["bb"]}</td><td>{rec["person"]}</td><td>{rec["tm"][-5:]}</td>'
        ls = rec['latest_status']
        lc = {'在用':'green','停机':'red','预拆机':'red'}.get(ls,'')
        html += f'<td><span class="tag {lc}">{ls}</span></td>'
        lt = str(rec['load_time'])[:16] if rec['load_time'] else '—'
        html += f'<td class="small">{lt}</td><td class="tr">{rec["value_pts"]}</td>'
        for i in range(14):
            s = rec[f's{i}']
            sc = {'在用':'green','停机':'red','预拆机':'red'}.get(s,'')
            html += f'<td><span class="tag {sc}" style="font-size:9px;padding:1px 4px">{s[:2] if s else "—"}</span></td>'
        html += '</tr>'
    html += '</tbody></table></div></div></div>'

# 图表数据
chart_data = {'tags': t_tags}
for p in projects_order:
    recs = projects.get(p, [])
    chart_data[p] = {
        'total': len(recs),
        'inuse': [sum(1 for r in recs if r[f's{i}'] == '在用') for i in range(14)],
        'stop': [sum(1 for r in recs if r[f's{i}'] == '停机') for i in range(14)],
    }

# 汇总数据
all_inuse = [sum(1 for r in records if r[f's{i}'] == '在用') for i in range(14)]
all_stop = [sum(1 for r in records if r[f's{i}'] == '停机') for i in range(14)]
all_pre = [sum(1 for r in records if r[f's{i}'] == '预拆机') for i in range(14)]

html += f'''</div>
<script>
var allInuse = {json.dumps(all_inuse)};
var allStop = {json.dumps(all_stop)};
var allPre = {json.dumps(all_pre)};
var labels = {json.dumps(t_tags)};

// 总览趋势图
new Chart(document.getElementById('trendChart'), {{
  type:'line',
  data:{{
    labels:labels,
    datasets:[
      {{label:'在用',data:allInuse,borderColor:'#2e7d32',backgroundColor:'rgba(46,125,50,.08)',fill:true,tension:.3,pointRadius:3}},
      {{label:'停机',data:allStop,borderColor:'#e65100',backgroundColor:'rgba(230,81,0,.08)',fill:true,tension:.3,pointRadius:3}},
      {{label:'预拆机',data:allPre,borderColor:'#c62828',backgroundColor:'rgba(198,40,40,.08)',fill:true,tension:.3,pointRadius:3}}
    ]}},
  options:{{
    responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{position:'top',labels:{{boxWidth:12,font:{{size:11}}}}}}}},
    scales:{{y:{{beginAtZero:true,grid:{{color:'rgba(0,0,0,.05)'}}}},x:{{grid:{{display:false}},ticks:{{font:{{size:10}}}}}}}}
  }}
}});

function filterProjTable(input, proj){{
  var q = input.value;
  document.querySelectorAll('#tbody-'+proj+' tr').forEach(function(r){{
    r.style.display = r.textContent.includes(q) ? '' : 'none';
  }});
}}
</script>
</body></html>'''

os.makedirs(os.path.join(PROJ, "docs"), exist_ok=True)
with open(OUT, 'w', encoding='utf-8') as f:
    f.write(html)
print(f"✅ 项目状态监控看板已生成: {OUT} ({len(html)/1024:.0f} KB)")
