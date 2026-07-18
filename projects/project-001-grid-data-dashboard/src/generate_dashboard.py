#!/usr/bin/env python3
"""
生成综合网格积分数据看板（4月、5月、6月 + 环比分析）
"""

import openpyxl, json, os, math
from collections import defaultdict

PROJ_DIR = "/Users/mr.g/Documents/Codex/Workspace/projects/project-001-grid-data-dashboard"

MONTHS = {
    "1月": {"file": "data/网格单元运营数据_202601.xlsx", "prefix": "01", "days": 31, "label": "1月（全月）"},
    "2月": {"file": "data/网格单元运营数据_202602.xlsx", "prefix": "02", "days": 28, "label": "2月（全月）"},
    "3月": {"file": "data/网格单元运营数据_202603.xlsx", "prefix": "03", "days": 31, "label": "3月（全月）"},
    "4月": {"file": "data/网格单元运营数据_202604.xlsx", "prefix": "04", "days": 30, "label": "4月（全月）"},
    "5月": {"file": "data/网格单元运营数据_202605.xlsx", "prefix": "05", "days": 31, "label": "5月（全月）"},
    "6月": {"file": "data/网格单元运营数据_202606.xlsx", "prefix": "06", "days": 30, "label": "6月（全月）"},
    "7月": {"file": "data/网格单元运营数据_202607.xlsx", "prefix": "07", "days": 0, "label": "7月（截至7月9日）"},
}

CATEGORIES = ["优惠到期", "存量变更", "拆机销户", "纯新套餐", "存量加装"]
CAT_COLORS = ["#f9a825", "#7b1fa2", "#c62828", "#2e7d32", "#00695c"]
CAT_BG = ["#fff3e0", "#f3e5f5", "#fce4ec", "#e8f5e9", "#e0f7fa"]

def safe_float(v):
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ('<null>', 'none', 'null', ''):
        return None
    try:
        return float(v)
    except:
        return None

def read_month_data(filepath):
    """Read daily data from Excel file, return structured data."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    headers = [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    
    col_map = {}
    date_col = None
    name_col = None
    
    for i, h in enumerate(headers, 1):
        if h == '网格单元名称':
            name_col = i
        elif h in ('统计日期', '计日期'):
            date_col = i
        elif h in CATEGORIES:
            col_map[i] = h
    
    # Group data by grid and by date
    grid_daily = defaultdict(list)  # grid_name -> list of {date, vals, total}
    date_set = set()
    
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, name_col).value or '').strip()
        if not name:
            continue
        
        date_str = str(ws.cell(r, date_col).value or '').strip()
        if not date_str or date_str in ('<null>', 'None', ''):
            continue
        date_set.add(date_str)
        
        vals = {}
        for ci, cn in col_map.items():
            vals[cn] = safe_float(ws.cell(r, ci).value)
        
        total = sum(v for v in vals.values() if v is not None)
        grid_daily[name].append({
            'date': date_str,
            'vals': vals,
            'total': round(total, 2)
        })
    
    wb.close()
    
    # Sort dates
    date_list = sorted(date_set)
    
    # Build per-date totals (dailyTotals)
    daily_totals = []
    for d in date_list:
        row_d = {'date': d}
        cat_sums = {c: 0 for c in CATEGORIES}
        total_sum = 0
        for name, entries in grid_daily.items():
            for entry in entries:
                if entry['date'] == d:
                    total_sum += entry['total']
                    for c in CATEGORIES:
                        v = entry['vals'].get(c)
                        if v is not None:
                            cat_sums[c] += v
        row_d.update(cat_sums)
        row_d['total'] = round(total_sum, 2)
        daily_totals.append(row_d)
    
    # Build full ranking (MTD totals per grid)
    grid_totals = {}
    for name, entries in grid_daily.items():
        totals_by_cat = {c: 0 for c in CATEGORIES}
        total_sum = 0
        for entry in entries:
            total_sum += entry['total']
            for c in CATEGORIES:
                v = entry['vals'].get(c)
                if v is not None:
                    totals_by_cat[c] += v
        for c in CATEGORIES:
            totals_by_cat[c] = round(totals_by_cat[c], 2)
        grid_totals[name] = {'total': round(total_sum, 2), **totals_by_cat}
    
    ranking = sorted(grid_totals.items(), key=lambda x: -x[1]['total'])
    total_grids = len(ranking)
    full_ranking = []
    for rank, (name, data) in enumerate(ranking, 1):
        full_ranking.append({
            'rank': rank,
            'name': name,
            'total': data['total'],
            **{c: data[c] for c in CATEGORIES}
        })
    
    top10 = full_ranking[:10]
    bot5 = full_ranking[-5:] if len(full_ranking) >= 5 else full_ranking
    
    pos_count = sum(1 for _, d in ranking if d['total'] > 0)
    neg_count = sum(1 for _, d in ranking if d['total'] < 0)
    zero_count = sum(1 for _, d in ranking if d['total'] == 0)
    sum_net = sum(d['total'] for _, d in ranking)
    avg_net = round(sum_net / total_grids, 2) if total_grids else 0
    
    # Category aggregates
    cat_stats = {}
    for c in CATEGORIES:
        vals = [d[c] for _, d in ranking]
        has_data = [v for v in vals if v is not None and v != 0]
        cat_sum = sum(v for v in vals if v is not None)
        cat_pos = sum(1 for v in vals if v is not None and v > 0)
        cat_neg = sum(1 for v in vals if v is not None and v < 0)
        cat_stats[c] = {'sum': round(cat_sum, 2), 'pos': cat_pos, 'neg': cat_neg}
    
    return {
        'date_list': date_list,
        'daily_totals': daily_totals,
        'grid_daily': dict(grid_daily),
        'full_ranking': full_ranking,
        'top10': top10,
        'bot5': bot5,
        'total_grids': total_grids,
        'sum_net': round(sum_net, 2),
        'avg_net': avg_net,
        'pos_count': pos_count,
        'neg_count': neg_count,
        'zero_count': zero_count,
        'cat_stats': cat_stats,
        'ranking_dict': {name: data for name, data in ranking},
    }

def json_serialize(obj):
    """Serialize data to JSON, handling None values properly."""
    if isinstance(obj, dict):
        return {k: json_serialize(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [json_serialize(v) for v in obj]
    elif isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return round(obj, 2)
    elif obj is None:
        return None
    return obj


# Read all months
print("读取数据中...")
all_data = {}
for month_key, m_info in MONTHS.items():
    filepath = os.path.join(PROJ_DIR, m_info['file'])
    print(f"  读取 {month_key}: {filepath}")
    all_data[month_key] = read_month_data(filepath)
    print(f"    网格: {all_data[month_key]['total_grids']}, 日期: {len(all_data[month_key]['date_list'])}, 净增合计: {all_data[month_key]['sum_net']}")

print("数据读取完毕，生成HTML...")

# Serialize all data to JSON for embedding
month_data_json = {}
for mk in MONTHS.keys():
    d = all_data[mk]
    # For grid_daily, serialize as {name: entries}
    gd_serialized = {}
    for name, entries in d['grid_daily'].items():
        gd_serialized[name] = [{'date': e['date'], 'vals': e['vals'], 'total': e['total']} for e in entries]
    month_data_json[mk] = {
        'dateList': d['date_list'],
        'dailyTotals': d['daily_totals'],
        'gh': json_serialize(gd_serialized),
        'fullRanking': d['full_ranking'],
        'totalGrids': d['total_grids'],
        'sumNet': d['sum_net'],
        'avgNet': d['avg_net'],
        'posCount': d['pos_count'],
        'negCount': d['neg_count'],
        'zeroCount': d['zero_count'],
        'catStats': d['cat_stats'],
    }

json_str = json.dumps(json_serialize(month_data_json), ensure_ascii=False, indent=None)

# 读取优化建议数据
recommendations = []
_rec_fp = os.path.join(PROJ_DIR, "data/recommendations.json")
if os.path.exists(_rec_fp):
    with open(_rec_fp, 'r', encoding='utf-8') as _f:
        _rec_data = json.load(_f)
    if isinstance(_rec_data, list):
        recommendations = _rec_data

# 构建优化建议 HTML
_rec_html = ''
if recommendations:
    _rec_html = '<div class="panel" style="border-top:none;"><div class="section-title" style="border-bottom:2px solid #e8eaf6;padding-bottom:6px;margin-bottom:8px;">\U0001f4a1 优化建议 <span style="font-size:12px;font-weight:400;color:#888;">（基于1-7月数据自动分析）</span></div><div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(360px,1fr));gap:8px;">'
    for r in recommendations:
        if r.get('flags'):
            _fls = ''
            for _f in r['flags']:
                if isinstance(_f, list) and len(_f) >= 2:
                    if '\u62c6\u673a' in str(_f[0]) or '\u6076\u5316' in str(_f[0]) or '\u98ce\u9669' in str(_f[0]):
                        _bg, _fg = '#fce4ec', '#c62828'
                    elif '\u65b0\u88c5' in str(_f[0]) or '\u5f3a\u52b2' in str(_f[0]) or '\u6539\u5584' in str(_f[0]):
                        _bg, _fg = '#e8f5e9', '#2e7d32'
                    else:
                        _bg, _fg = '#fff3e0', '#e65100'
                    _fls += '<div style="margin:2px 0;line-height:1.5;"><span style="display:inline-block;padding:1px 8px;border-radius:3px;font-size:10px;font-weight:500;background:' + _bg + ';color:' + _fg + ';margin-right:6px;">' + str(_f[0]) + '</span><span style="font-size:11px;color:#555;">' + str(_f[1]) + '</span></div>'
            if _fls:
                _rec_html += '<div style="padding:8px 10px;background:#fafbff;border-radius:8px;border:1px solid #f0f0f0;"><b style="font-size:12px;color:#1a237e;">' + str(r['grid']) + '</b>' + _fls + '</div>'
    _rec_html += '</div></div>'


# ---- Build HTML ----
html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>网格积分数据看板（2026年1-7月）</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+SC:wght@400;500;600;700&display=swap');
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Noto Sans SC',sans-serif;background:#f0f2f5;color:#1a1a2e}}
.topbar{{background:linear-gradient(135deg,#0d1b4a,#1a237e,#283593);color:#fff;padding:14px 28px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:6px}}
.topbar h1{{font-size:17px;letter-spacing:1px;font-weight:700}}
.topbar .meta{{font-size:12px;opacity:.7;display:flex;gap:10px;flex-wrap:wrap}}
.dashboard{{max-width:1200px;margin:0 auto;padding:12px 16px}}
.kpi-section{{background:#fff;border-radius:10px;box-shadow:0 1px 6px rgba(0,0,0,.04);padding:14px 18px;margin-bottom:14px}}
.kpi-top{{display:flex;gap:10px;margin-bottom:10px;flex-wrap:wrap}}
.kpi-main{{flex:1;display:grid;grid-template-columns:repeat(4,1fr);gap:8px}}
.kpi-item{{text-align:center;padding:8px 6px;border-radius:6px}}
.kpi-item .kv{{font-size:20px;font-weight:700;font-family:'Menlo',monospace}}
.kpi-item .kl{{font-size:11px;color:#666;margin-top:1px}}
.kpi-breakdown{{display:grid;grid-template-columns:repeat(5,1fr);gap:6px}}
.kpi-sub{{padding:6px 8px;border-radius:6px;text-align:center;font-size:12px;display:flex;flex-direction:column;align-items:center}}
.kpi-sub .sv{{font-family:'Menlo',monospace;font-weight:600;font-size:15px}}
.kpi-sub .sl{{font-size:11px;color:#666;margin-top:1px}}
.month-tabs{{display:flex;gap:2px;background:#e8eaf6;border-radius:8px;padding:3px;margin-bottom:10px;overflow-x:auto}}
.mtab-btn{{padding:8px 14px;border:none;background:transparent;border-radius:6px;font-size:13px;font-weight:600;cursor:pointer;color:#666;font-family:inherit;white-space:nowrap;transition:all .15s}}
.mtab-btn:hover{{background:rgba(255,255,255,.4);color:#1a237e}}
.mtab-btn.active{{background:#fff;color:#1a237e;box-shadow:0 1px 4px rgba(0,0,0,.08)}}
.mtab-panel{{display:none}}.mtab-panel.active{{display:block}}
.sub-tabs{{display:flex;gap:2px;background:#e8eaf6;border-radius:8px;padding:3px;margin-bottom:14px;overflow-x:auto}}
.sub-tab-btn{{padding:8px 16px;border:none;background:transparent;border-radius:6px;font-size:13px;font-weight:600;cursor:pointer;color:#666;font-family:inherit;white-space:nowrap}}
.sub-tab-btn:hover{{background:rgba(255,255,255,.4);color:#1a237e}}
.sub-tab-btn.active{{background:#fff;color:#1a237e;box-shadow:0 1px 4px rgba(0,0,0,.08)}}
.sub-tab-panel{{display:none}}.sub-tab-panel.active{{display:block}}
.panel{{background:#fff;border-radius:10px;box-shadow:0 1px 6px rgba(0,0,0,.04);padding:16px;margin-bottom:14px}}
.panel-title{{font-size:14px;font-weight:700;color:#1a237e;padding-bottom:8px;border-bottom:2px solid #e8eaf6;margin-bottom:12px;display:flex;align-items:center;gap:6px}}
.chart-row{{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:14px}}
.chart-box{{background:#fff;border-radius:10px;padding:16px;box-shadow:0 1px 6px rgba(0,0,0,.04)}}
.chart-box h3{{font-size:13px;color:#1a237e;margin-bottom:8px;font-weight:600}}
.chart-wrap{{width:100%;height:260px;position:relative}}
.dist-bar{{display:flex;height:22px;border-radius:11px;overflow:hidden;margin:8px 0}}
.dist-legend{{display:flex;flex-wrap:wrap;gap:5px;margin-top:6px;justify-content:center}}
.dist-legend span{{font-size:11px;color:#555;display:flex;align-items:center;gap:3px}}
.dist-dot{{width:8px;height:8px;border-radius:50%;display:inline-block}}
.tbl-wrap{{overflow-x:auto;border-radius:6px;border:1px solid #eee}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
th{{background:#e8eaf6;color:#1a237e;font-weight:600;padding:6px 8px;text-align:center;font-size:11px;position:sticky;top:0;z-index:1}}
td{{padding:5px 8px;border-bottom:1px solid #f0f0f0;text-align:center}}
tr:hover td{{background:#f8f9ff}}
.glink{{cursor:pointer;color:#1a237e;font-weight:500}}
.glink:hover{{text-decoration:underline;background:#e8eaf6!important;border-radius:3px}}
.tl{{text-align:left}}
.tr{{text-align:right;font-variant-numeric:tabular-nums;font-family:'Menlo',monospace;font-size:12px}}
.p{{color:#2e7d32}}
.n{{color:#c62828}}
.scroll-y{{max-height:600px;overflow-y:auto;border-radius:6px;border:1px solid #eee}}
.dp{{display:flex;flex-wrap:wrap;gap:4px;margin-bottom:10px}}
.dp-btn{{padding:4px 10px;border:1px solid #ddd;border-radius:5px;background:#fff;cursor:pointer;font-size:12px;font-family:inherit;color:#555}}
.dp-btn:hover{{border-color:#3949ab;color:#1a237e}}
.dp-btn.active{{background:#1a237e;color:#fff;border-color:#1a237e}}
.g2{{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:14px}}
.g3{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin-bottom:14px}}
.footer{{text-align:center;font-size:11px;color:#bbb;margin:16px 0 8px}}
.modal-overlay{{display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,.4);z-index:999;align-items:center;justify-content:center}}
.modal-overlay.show{{display:flex}}
.modal-box{{background:#fff;border-radius:12px;padding:0;max-width:800px;width:90vw;max-height:85vh;display:flex;flex-direction:column;box-shadow:0 8px 40px rgba(0,0,0,.2)}}
.modal-header{{display:flex;justify-content:space-between;align-items:center;padding:14px 20px;border-bottom:1px solid #eee}}
.modal-header h2{{font-size:16px;color:#1a237e;font-weight:700}}
.modal-close{{width:30px;height:30px;border:none;background:#f0f0f0;border-radius:6px;font-size:18px;cursor:pointer;display:flex;align-items:center;justify-content:center;color:#666;font-family:inherit}}
.modal-close:hover{{background:#e0e0e0;color:#333}}
.modal-body{{padding:16px 20px;overflow-y:auto;flex:1}}
.modal-summary{{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:12px}}
.modal-summary .ms-item{{text-align:center;padding:8px;border-radius:6px;font-size:12px}}
.modal-summary .ms-item .msv{{font-family:'Menlo',monospace;font-weight:700;font-size:16px}}
.modal-body table{{font-size:12px}}
.modal-body th{{font-size:10px;padding:5px 6px}}
.modal-body td{{padding:4px 6px;font-size:12px}}
@media(max-width:768px){{.kpi-main{{grid-template-columns:repeat(2,1fr)}}.kpi-breakdown{{grid-template-columns:repeat(3,1fr)}}.chart-row,.g2,.g3{{grid-template-columns:1fr}}.topbar{{padding:10px}}.dashboard{{padding:6px}}.modal-box{{width:95vw;max-height:90vh}}}}

/* 环比分析样式 */
.hb-summary{{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:10px;margin-bottom:14px}}
.hb-card{{background:#fff;border-radius:10px;padding:14px;box-shadow:0 1px 6px rgba(0,0,0,.04);text-align:center}}
.hb-card h3{{font-size:13px;color:#1a237e;margin-bottom:4px;font-weight:600}}
.hb-card .hl{{font-size:11px;color:#888;margin-bottom:2px}}
.hb-card .hv{{font-size:22px;font-weight:700;font-family:'Menlo',monospace}}
.hb-card .hv2{{font-size:16px;font-weight:600;font-family:'Menlo',monospace;margin-top:2px}}
.hb-trend{{font-size:12px;margin-top:4px}}
.hb-up{{color:#2e7d32}}
.hb-down{{color:#c62828}}
.hb-eq{{color:#888}}
.hb-detail-grid{{display:grid;grid-template-columns:1fr 1fr;gap:12px}}
.hb-comp-card{{margin-bottom:10px;border-radius:8px;padding:10px 12px;background:#fff;border:1px solid #eee}}
.hb-comp-header{{display:flex;justify-content:space-between;align-items:center;margin-bottom:4px}}
.hb-comp-name{{font-weight:600;font-size:13px}}
.hb-comp-delta{{font-family:'Menlo',monospace;font-size:13px;font-weight:600}}
.hb-comp-sub{{font-size:11px;color:#888;display:flex;gap:10px}}
.gb-comp{{display:grid;grid-template-columns:1fr 1fr;gap:8px}}
.comp-table{{font-size:12px}}
.comp-table td{{font-size:11px;padding:4px 6px}}
.comp-table th{{font-size:10px;padding:4px 6px}}

/* 环比子标签 */
.hb-sub-tabs{{display:flex;gap:2px;background:#e8eaf6;border-radius:8px;padding:3px;margin-bottom:12px;overflow-x:auto}}
.hb-tab-btn{{padding:8px 14px;border:none;background:transparent;border-radius:6px;font-size:13px;font-weight:600;cursor:pointer;color:#666;font-family:inherit;white-space:nowrap}}
.hb-tab-btn:hover{{background:rgba(255,255,255,.4);color:#1a237e}}
.hb-tab-btn.active{{background:#fff;color:#1a237e;box-shadow:0 1px 4px rgba(0,0,0,.08)}}
.hb-panel{{display:none}}.hb-panel.active{{display:block}}

.formula-row{{display:flex;align-items:center;gap:4px;flex-wrap:wrap}}
.formula-row .fv{{font-family:'Menlo',monospace;font-weight:600;font-size:13px}}
.formula-row .arrow{{color:#888;font-size:11px}}
</style>
</head>
<body>
<div class="topbar">
  <h1>📊 网格积分数据看板</h1>
  <div class="meta">
    <span>📅 2026年1月 - 7月</span>
    <span id="topGridCount"></span>
    <span id="topDateRange"></span>
  </div>
</div>
<div class="dashboard">

<div class="month-tabs" id="monthTabs"></div>

<div class="mtab-panel active" id="mp-1月"></div>
  <div class="mtab-panel" id="mp-2月"></div>
  <div class="mtab-panel" id="mp-3月"></div>
  <div class="mtab-panel" id="mp-4月"></div>
<div class="mtab-panel" id="mp-5月"></div>
<div class="mtab-panel" id="mp-6月"></div>
  <div class="mtab-panel" id="mp-7月"></div>
<div class="mtab-panel" id="mp-环比分析"></div>
  <div class="mtab-panel" id="mp-优化建议">{_rec_html}</div>

</div>

<div class="modal-overlay" id="modal">
  <div class="modal-box">
    <div class="modal-header">
      <h2 id="modalTitle">网格明细</h2>
      <button class="modal-close" onclick="document.getElementById('modal').classList.remove('show')">✕</button>
    </div>
    <div class="modal-body" id="modalBody"></div>
  </div>
</div>

<script>
var MONTH_DATA = {json_str};
var CATEGORIES = {json.dumps(CATEGORIES, ensure_ascii=False)};
var CAT_COLORS = {json.dumps(CAT_COLORS, ensure_ascii=False)};
var CAT_BG = {json.dumps(CAT_BG, ensure_ascii=False)};
var monthKeys = ['1月','2月','3月','4月','5月','6月','7月'];
var charts = {{}};

function monthDir(mk) {{
  if (mk === '4月') return '04';
  if (mk === '5月') return '05';
  if (mk === '6月') return '06';
  return '';
}}

function showGrid(mk, name) {{
  var data = MONTH_DATA[mk].gh[name];
  if (!data || data.length === 0) {{ alert('暂无该网格的每日数据'); return; }}
  data.sort(function(a, b) {{ return a.date.localeCompare(b.date); }});
  document.getElementById('modalTitle').textContent = '📋 ' + name + ' - ' + mk;
  var total = 0, posDays = 0, negDays = 0;
  var ct = {{}}; CATEGORIES.forEach(function(c) {{ ct[c] = 0; }});
  data.forEach(function(d) {{
    total += d.total;
    if (d.total > 0) posDays++; else if (d.total < 0) negDays++;
    CATEGORIES.forEach(function(c) {{ if (d.vals[c]) ct[c] += d.vals[c]; }});
  }});
  var h = '<div class="modal-summary">' +
    '<div class="ms-item" style="background:#e3f2fd;"><div class="msv" style="color:#1565c0;">' + total.toFixed(2) + '</div><div>MTD净增</div></div>' +
    '<div class="ms-item" style="background:#e8f5e9;"><div class="msv" style="color:#2e7d32;">' + posDays + '</div><div>正增天数</div></div>' +
    '<div class="ms-item" style="background:#fce4ec;"><div class="msv" style="color:#c62828;">' + negDays + '</div><div>负增天数</div></div></div>';
  h += '<table><thead><tr><th>日期</th><th>净增</th>';
  CATEGORIES.forEach(function(c) {{ h += '<th>' + c + '</th>'; }});
  h += '</tr></thead><tbody>';
  data.forEach(function(d) {{
    var cls = d.total >= 0 ? 'p' : 'n';
    var vs = CATEGORIES.map(function(c) {{
      var v = d.vals[c];
      if (v === undefined || v === null) return '<td class="tr" style="opacity:.3">—</td>';
      return '<td class="tr ' + (v > 0 ? 'p' : v < 0 ? 'n' : '') + '">' + v.toFixed(2) + '</td>';
    }}).join('');
    h += '<tr><td class="tl">' + d.date.slice(4,6) + '-' + d.date.slice(6,8) + '</td><td class="tr ' + cls + '">' + d.total.toFixed(2) + '</td>' + vs + '</tr>';
  }});
  h += '</tbody></table>';
  document.getElementById('modalBody').innerHTML = h;
  document.getElementById('modal').classList.add('show');
}}


function filterGridTable(input) {{
  var filter = input.value.toLowerCase().trim();
  var panel = input.closest('.panel');
  if (!panel) return;
  var scrollDiv = panel.querySelector('.scroll-y');
  if (!scrollDiv) return;
  var table = scrollDiv.querySelector('table');
  if (!table) return;
  var rows = table.querySelectorAll('tbody tr');
  rows.forEach(function(row) {{
    var nameCell = row.querySelector('td:nth-child(2)');
    if (nameCell) {{
      var text = nameCell.textContent.toLowerCase();
      row.style.display = text.indexOf(filter) > -1 ? '' : 'none';
    }}
  }});
}}

function renderMonth(mk) {{
  var MD = MONTH_DATA[mk];
  if (!MD) return;
  var dp = monthDir(mk);
  var dates = MD.dateList;
  var dt = MD.dailyTotals;
  var fr = MD.fullRanking;
  var prefix = mk;

  // ---- KPI ----
  var kpiBG = ['#e3f2fd','#e8f5e9','#fce4ec','#fff3e0'];
  var kpiColor = ['#1565c0','#2e7d32','#c62828','#e65100'];
  var kpiLabels = ['📊 净增合计（MTD）','✅ 正净增网格','❌ 负净增网格','📈 平均净增'];
  var kpiVals = [MD.sumNet, MD.posCount, MD.negCount, MD.avgNet];
  var kpiFmt = function(v,i) {{ return i === 0 || i === 3 ? v.toFixed(2) : v.toString(); }};

  var h = '';
  h += '<div class="kpi-section">';
  h += '<div class="kpi-top"><div class="kpi-main">';
  kpiLabels.forEach(function(lb,i) {{
    h += '<div class="kpi-item" style="background:' + kpiBG[i] + ';">';
    h += '<div class="kv" style="color:' + kpiColor[i] + ';">';
    if (i === 0 || i === 3) {{
      h += MD.sumNet >= 0 ? '+' : '';
      h += kpiFmt(kpiVals[i], i);
    }} else {{
      h += kpiFmt(kpiVals[i], i);
    }}
    h += '</div><div class="kl">' + lb + '</div></div>';
  }});
  h += '</div></div>';
  h += '<div class="kpi-breakdown">';
  CATEGORIES.forEach(function(c, i) {{
    var s = MD.catStats[c] || {{sum:0}};
    var cls = s.sum >= 0 ? 'p' : 'n';
    h += '<div class="kpi-sub" style="background:' + CAT_BG[i] + ';">';
    h += '<span class="sv" style="color:' + CAT_COLORS[i] + ';">' + (s.sum >= 0 ? '+' : '') + s.sum.toFixed(2) + '</span>';
    h += '<span class="sl">' + c + '</span></div>';
  }});
  h += '</div></div>';

  // ---- Sub-tabs ----
  h += '<div class="sub-tabs" id="subTabs-' + mk + '">';
  h += '<button class="sub-tab-btn active" data-sub="' + mk + '-overview">📊 概览</button>';
  h += '<button class="sub-tab-btn" data-sub="' + mk + '-daily">📅 每日明细</button>';
  h += '<button class="sub-tab-btn" data-sub="' + mk + '-rank">🏆 排名</button>';
  h += '</div>';

  // ---- Overview tab ----
  h += '<div class="sub-tab-panel active" id="st-' + mk + '-overview">';
  
  // Trend charts (only for months with multiple days)
  if (dates.length > 1) {{
    h += '<div class="chart-row"><div class="chart-box"><h3>📈 每日净增趋势</h3><div class="chart-wrap"><canvas id="netChart-' + mk + '"></canvas></div></div>';
    h += '<div class="chart-box"><h3>📊 累计净增曲线</h3><div class="chart-wrap"><canvas id="cumChart-' + mk + '"></canvas></div></div></div>';
  }}
  
  // Category breakdown chart
  h += '<div class="chart-row"><div class="chart-box"><h3>📊 各类别占比</h3><div class="chart-wrap"><canvas id="catChart-' + mk + '"></canvas></div></div>';
  h += '<div class="chart-box"><h3>📈 各网格净增分布</h3><div class="chart-wrap"><canvas id="distChart-' + mk + '"></canvas></div></div></div>';

  // Daily summary table
  h += '<div class="panel"><div class="panel-title">📅 ' + mk + '每日汇总</div><div class="tbl-wrap"><table><thead><tr><th>日期</th><th>合计</th>';
  CATEGORIES.forEach(function(c) {{ h += '<th>' + c + '</th>'; }});
  h += '</tr></thead><tbody>';
  dt.forEach(function(d) {{
    var dateLabel = d.date.slice(4,6) + '-' + d.date.slice(6,8);
    var cls = d.total >= 0 ? 'p' : 'n';
    h += '<tr><td class="tl">' + dateLabel + '</td><td class="tr ' + cls + '">' + d.total.toFixed(2) + '</td>';
    CATEGORIES.forEach(function(c) {{
      var v = d[c] || 0;
      var cc = v > 0 ? 'p' : v < 0 ? 'n' : '';
      h += '<td class="tr ' + cc + '">' + v.toFixed(2) + '</td>';
    }});
    h += '</tr>';
  }});
  h += '</tbody></table></div></div>';
  h += '</div>'; // end overview

  // ---- Daily tab ----
  h += '<div class="sub-tab-panel" id="st-' + mk + '-daily">';
  h += '<div class="panel"><div class="panel-title">📅 每日积分发展情况</div>';
  h += '<div class="dp" id="datePicker-' + mk + '"></div>';
  h += '<div class="tbl-wrap" id="dailyTable-' + mk + '"></div></div></div>';

  // ---- Rank tab ----
  h += '<div class="sub-tab-panel" id="st-' + mk + '-rank">';
  h += '<div class="g2"><div class="panel"><div class="panel-title" style="color:#2e7d32;">🏆 MTD TOP 10</div><div class="tbl-wrap"><table><thead><tr><th style="width:40px">#</th><th class="tl">网格名称</th><th>净增</th><th>纯新套餐</th><th>拆机销户</th></tr></thead><tbody>';
  var t10 = MD.fullRanking.slice(0,10);
  t10.forEach(function(r) {{
    h += '<tr><td>' + r.rank + '</td><td class="tl glink" onclick="showGrid(\\'' + mk + '\\',\\'' + r.name.replace(/'/g,"\\\\'") + '\\')">' + r.name + '</td>';
    h += '<td class="tr ' + (r.total >= 0 ? 'p' : 'n') + '">' + r.total.toFixed(2) + '</td>';
    h += '<td class="tr ' + (r['纯新套餐'] >= 0 ? 'p' : 'n') + '">' + (r['纯新套餐'] || 0).toFixed(2) + '</td>';
    h += '<td class="tr ' + (r['拆机销户'] >= 0 ? 'p' : 'n') + '">' + (r['拆机销户'] || 0).toFixed(2) + '</td></tr>';
  }});
  h += '</tbody></table></div></div>';

  h += '<div class="panel"><div class="panel-title" style="color:#c62828;">⚠️ MTD BOTTOM 5</div><div class="tbl-wrap"><table><thead><tr><th style="width:40px">#</th><th class="tl">网格名称</th><th>净增</th><th>纯新套餐</th><th>拆机销户</th></tr></thead><tbody>';
  var b5 = MD.fullRanking.slice(-5);
  b5.forEach(function(r) {{
    h += '<tr><td>' + r.rank + '</td><td class="tl glink" onclick="showGrid(\\'' + mk + '\\',\\'' + r.name.replace(/'/g,"\\\\'") + '\\')">' + r.name + '</td>';
    h += '<td class="tr n">' + r.total.toFixed(2) + '</td>';
    h += '<td class="tr ' + (r['纯新套餐'] >= 0 ? 'p' : 'n') + '">' + (r['纯新套餐'] || 0).toFixed(2) + '</td>';
    h += '<td class="tr ' + (r['拆机销户'] >= 0 ? 'p' : 'n') + '">' + (r['拆机销户'] || 0).toFixed(2) + '</td></tr>';
  }});
  h += '</tbody></table></div></div></div>';

 h += '<div class="panel"><div class="panel-title">📋 全部网格 MTD 排名 <span style="font-weight:400;font-size:12px;color:#888;">（点击网格名查看每日明细）</span></div>';
  h += '<div style="margin-bottom:8px;"><input type="text" placeholder="🔍 搜索网格名称..." oninput="filterGridTable(this)" style="width:100%;padding:8px 12px;border:1px solid #ddd;border-radius:6px;font-size:13px;font-family:inherit;outline:none;box-sizing:border-box;"></div>';
 h += '<div class="scroll-y"><table><thead><tr><th style="width:40px">#</th><th class="tl">网格名称</th><th>净增积分</th>';
  CATEGORIES.forEach(function(c) {{ h += '<th>' + c + '</th>'; }});
  h += '</tr></thead><tbody>';
  fr.forEach(function(r) {{
    h += '<tr><td>' + r.rank + '</td><td class="tl glink" onclick="showGrid(\\'' + mk + '\\',\\'' + r.name.replace(/'/g,"\\\\'") + '\\')" style="cursor:pointer;color:#1a237e;font-weight:500;">' + r.name + '</td>';
    h += '<td class="tr ' + (r.total >= 0 ? 'p' : 'n') + '">' + r.total.toFixed(2) + '</td>';
    CATEGORIES.forEach(function(c) {{
      var v = r[c] || 0;
      var cc = v > 0 ? 'p' : v < 0 ? 'n' : '';
      h += '<td class="tr ' + cc + '">' + v.toFixed(2) + '</td>';
    }});
    h += '</tr>';
  }});
  h += '</tbody></table></div></div>';
  h += '</div>'; // end rank tab

  document.getElementById('mp-' + mk).innerHTML = h;

  // ---- Boot charts ----
  setTimeout(function() {{
    bootMonthCharts(mk);
    bootDailyPicker(mk);
  }}, 50);
}}

function bootMonthCharts(mk) {{
  var MD = MONTH_DATA[mk];
  var dates = MD.dateList;
  var dt = MD.dailyTotals;
  var fr = MD.fullRanking;
  var ctx;

  // Net trend chart
  var netCanv = document.getElementById('netChart-' + mk);
  if (netCanv && dates.length > 1) {{
    ctx = netCanv.getContext('2d');
    if (charts['net-' + mk]) charts['net-' + mk].destroy();
    var labels = dates.map(function(d) {{ return d.slice(4,6) + '-' + d.slice(6,8); }});
    var netData = dt.map(function(d) {{ return d.total; }});
    var posData = dt.map(function(d) {{ return Math.max(0, d.total); }});
    var negData = dt.map(function(d) {{ return Math.min(0, d.total); }});
    charts['net-' + mk] = new Chart(ctx, {{
      type: 'bar',
      data: {{
        labels: labels,
        datasets: [
          {{ label: '正净增', data: posData, backgroundColor: 'rgba(46,125,50,0.7)', borderRadius: 2, barPercentage: 0.6 }},
          {{ label: '负净增', data: negData, backgroundColor: 'rgba(198,40,40,0.7)', borderRadius: 2, barPercentage: 0.6 }}
        ]
      }},
      options: {{
        responsive: true, maintainAspectRatio: false,
        plugins: {{ legend: {{ display: true, position: 'top', labels: {{ boxWidth: 12, font: {{ size: 11 }} }} }} }},
        scales: {{ y: {{ grid: {{ color: 'rgba(0,0,0,0.05)' }}, ticks: {{ font: {{ size: 10 }} }} }}, x: {{ ticks: {{ font: {{ size: 10 }}, maxRotation: 45 }} }} }}
      }}
    }});
  }}

  // Cumulative chart
  var cumCanv = document.getElementById('cumChart-' + mk);
  if (cumCanv && dates.length > 1) {{
    ctx = cumCanv.getContext('2d');
    if (charts['cum-' + mk]) charts['cum-' + mk].destroy();
    var labels = dates.map(function(d) {{ return d.slice(4,6) + '-' + d.slice(6,8); }});
    var cum = 0;
    var cumData = dt.map(function(d) {{ cum += d.total; return Math.round(cum*100)/100; }});
    charts['cum-' + mk] = new Chart(ctx, {{
      type: 'line',
      data: {{
        labels: labels,
        datasets: [{{ label: '累计净增', data: cumData, borderColor: '#1a237e', backgroundColor: 'rgba(26,35,126,0.1)', fill: true, tension: 0.3, pointRadius: 3 }}]
      }},
      options: {{
        responsive: true, maintainAspectRatio: false,
        plugins: {{ legend: {{ display: true, position: 'top', labels: {{ boxWidth: 12, font: {{ size: 11 }} }} }} }},
        scales: {{ y: {{ grid: {{ color: 'rgba(0,0,0,0.05)' }}, ticks: {{ font: {{ size: 10 }} }} }}, x: {{ ticks: {{ font: {{ size: 10 }}, maxRotation: 45 }} }} }}
      }}
    }});
  }}

  // Category pie chart
  var catCanv = document.getElementById('catChart-' + mk);
  if (catCanv) {{
    ctx = catCanv.getContext('2d');
    if (charts['cat-' + mk]) charts['cat-' + mk].destroy();
    var catSums = CATEGORIES.map(function(c) {{ return Math.abs((MD.catStats[c] || {{sum:0}}).sum); }});
    charts['cat-' + mk] = new Chart(ctx, {{
      type: 'doughnut',
      data: {{
        labels: CATEGORIES,
        datasets: [{{ data: catSums, backgroundColor: CAT_COLORS.map(function(c){{return c+'CC'}}), borderWidth: 1, borderColor: '#fff' }}]
      }},
      options: {{
        responsive: true, maintainAspectRatio: false,
        plugins: {{ legend: {{ position: 'right', labels: {{ boxWidth: 12, font: {{ size: 11 }} }} }} }}
      }}
    }});
  }}

  // Distribution histogram
  var distCanv = document.getElementById('distChart-' + mk);
  if (distCanv) {{
    ctx = distCanv.getContext('2d');
    if (charts['dist-' + mk]) charts['dist-' + mk].destroy();
    var netVals = fr.map(function(r) {{ return r.total; }}).sort(function(a,b){{return a-b}});
    // Create bins
    var min = Math.floor(Math.min.apply(null, netVals));
    var max = Math.ceil(Math.max.apply(null, netVals));
    var binCount = 20;
    var binWidth = (max - min) / binCount || 1;
    var bins = new Array(binCount).fill(0);
    var binLabels = [];
    for (var i = 0; i < binCount; i++) {{
      var lo = min + i * binWidth;
      var hi = lo + binWidth;
      binLabels.push((lo|0) + '~' + (hi|0));
    }}
    netVals.forEach(function(v) {{
      var idx = Math.min(Math.floor((v - min) / binWidth), binCount-1);
      if (idx >= 0) bins[idx]++;
    }});
    charts['dist-' + mk] = new Chart(ctx, {{
      type: 'bar',
      data: {{
        labels: binLabels,
        datasets: [{{ label: '网格数', data: bins, backgroundColor: 'rgba(26,35,126,0.6)', borderRadius: 2 }}]
      }},
      options: {{
        responsive: true, maintainAspectRatio: false,
        plugins: {{ legend: {{ display: false }} }},
        scales: {{ y: {{ ticks: {{ font: {{ size: 10 }}, stepSize: 1 }} }}, x: {{ ticks: {{ font: {{ size: 8 }}, maxRotation: 60 }} }} }}
      }}
    }});
  }}
}}

var dailyRenderers = {{}};

function bootDailyPicker(mk) {{
  var MD = MONTH_DATA[mk];
  var dates = MD.dateList;
  if (dates.length === 0) return;
  var container = document.getElementById('datePicker-' + mk);
  var tableContainer = document.getElementById('dailyTable-' + mk);
  if (!container || !tableContainer) return;
  
  function renderDailyTable(date) {{
    var dtData = MD.dailyTotals.find(function(d) {{ return d.date === date; }});
    if (!dtData) {{ tableContainer.innerHTML = '<div style="padding:20px;text-align:center;color:#888">当日无数据</div>'; return; }}
    var grids = [];
    for (var name in MD.gh) {{
      MD.gh[name].forEach(function(e) {{
        if (e.date === date) {{
          grids.push({{ name: name, total: e.total, vals: e.vals }});
        }}
      }});
    }}
    grids.sort(function(a,b){{return b.total-a.total}});
    var h = '<table><thead><tr><th style="width:40px">#</th><th class="tl">网格名称</th><th>净增</th>';
    CATEGORIES.forEach(function(c) {{ h += '<th>' + c + '</th>'; }});
    h += '</tr></thead><tbody>';
    if (grids.length === 0) {{
      h += '<tr><td colspan="' + (3 + CATEGORIES.length) + '" style="padding:20px;text-align:center;color:#888">当日无数据</td></tr>';
    }} else {{
      grids.forEach(function(g, idx) {{
        h += '<tr><td>' + (idx+1) + '</td><td class="tl glink" onclick="showGrid(\\'' + mk + '\\',\\'' + g.name.replace(/'/g,"\\\\'") + '\\')">' + g.name + '</td>';
        h += '<td class="tr ' + (g.total >= 0 ? 'p' : 'n') + '">' + g.total.toFixed(2) + '</td>';
        CATEGORIES.forEach(function(c) {{
          var v = g.vals[c];
          if (v === undefined || v === null) {{ h += '<td class="tr" style="opacity:.3">—</td>'; }}
          else {{ h += '<td class="tr ' + (v > 0 ? 'p' : v < 0 ? 'n' : '') + '">' + v.toFixed(2) + '</td>'; }}
        }});
        h += '</tr>';
      }});
    }}
    h += '</tbody></table>';
    tableContainer.innerHTML = h;
  }}
  
  dailyRenderers[mk] = {{
    render: renderDailyTable,
    dates: dates,
    activeDate: dates[0]
  }};
  
  function buildDatePicker() {{
    var h = '';
    dates.forEach(function(d) {{
      var label = d.slice(4,6) + '-' + d.slice(6,8);
      var cls = d === dailyRenderers[mk].activeDate ? 'dp-btn active' : 'dp-btn';
      h += '<button class="' + cls + '" onclick="pickDate(\\'' + mk + '\\',\\'' + d + '\\')">' + label + '</button>';
    }});
    container.innerHTML = h;
    renderDailyTable(dates[0]);
  }}
  
  buildDatePicker();
}}

window.pickDate = function(mk2, date) {{
  var r = dailyRenderers[mk2];
  if (!r) return;
  r.activeDate = date;
  var btns = document.querySelectorAll('#datePicker-' + mk2 + ' .dp-btn');
  btns.forEach(function(b) {{ b.classList.remove('active'); }});
  var target = Array.from(btns).find(function(b) {{ return b.textContent.trim() === date.slice(4,6) + '-' + date.slice(6,8); }});
  if (target) target.classList.add('active');
  r.render(date);
}};

// ---- 环比分析 ----
function renderMonthComparison() {{
  var h = '';
  
  // Overview cards
  var mks = ['1月','2月','3月','4月','5月','6月','7月'];
  var sums = mks.map(function(mk) {{ return MONTH_DATA[mk].sumNet; }});
  var avgs = mks.map(function(mk) {{ return MONTH_DATA[mk].avgNet; }});
  var pos = mks.map(function(mk) {{ return MONTH_DATA[mk].posCount; }});
  var neg = mks.map(function(mk) {{ return MONTH_DATA[mk].negCount; }});
  var grids = mks.map(function(mk) {{ return MONTH_DATA[mk].totalGrids; }});
  
  // Net sum month over month
  var mks = ['1月','2月','3月','4月','5月','6月','7月'];
  var mData = {{}}, sums = [], dayAvgs = [], dayCnt = [31,28,31,30,31,30,5];
  mks.forEach(function(mk,i){{
    mData[mk] = MONTH_DATA[mk];
    sums.push(mData[mk].sumNet);
    dayAvgs.push(mData[mk].sumNet / dayCnt[i]);
  }});
  var chgPairs = [];
  for (var ci = 1; ci < mks.length; ci++) {{
    var prev = sums[ci-1], curr = sums[ci];
    var chg = curr - prev;
    var pct = prev !== 0 ? ((chg / Math.abs(prev)) * 100).toFixed(1) : 'N/A';
    chgPairs.push({{from: mks[ci-1], to: mks[ci], chg: chg, pct: pct}});
  }}

  // Tabs
  h += '<div class="hb-sub-tabs">';
  h += '<button class="hb-tab-btn active" data-hb="hb-overview">📊 概览</button>';
  h += '<button class="hb-tab-btn" data-hb="hb-daily">📈 日度趋势对比</button>';
  h += '<button class="hb-tab-btn" data-hb="hb-grids">🏘️ 网格环比</button>';
  h += '<button class="hb-tab-btn" data-hb="hb-detail">📋 排名变化</button>';
  h += '</div>';

  // Panel 1: Overview
  h += '<div class="hb-panel active" id="hbp-hb-overview">';
  
  h += '<div class="hb-summary">';
  var dayLabels = ['31天','28天','31天','30天','31天','30天','5天'];
  mks.forEach(function(mk,i){{
    h += '<div class="hb-card"><h3>' + mk + '<br><small>' + dayLabels[i] + '</small></h3>';
    h += '<div class="hl">净增合计（MTD）</div>';
    h += '<div class="hv" style="color:' + (sums[i] >= 0 ? '#2e7d32' : '#c62828') + ';">' + (sums[i] >= 0 ? '+' : '') + sums[i].toFixed(2) + '</div>';
    h += '<div class="hl" style="margin-top:6px;">日均净增</div>';
    h += '<div class="hv2" style="color:' + (dayAvgs[i] >= 0 ? '#2e7d32' : '#c62828') + ';">' + (dayAvgs[i] >= 0 ? '+' : '') + dayAvgs[i].toFixed(2) + '</div>';
    h += '<div class="hl" style="margin-top:6px;">正/负净增网格</div>';
    h += '<div class="hv2" style="font-size:14px;">👍' + pos[i] + ' 👎' + neg[i] + '</div>';
    h += '</div>';
  }});
  h += '</div>';

  // Category comparison table
  h += '<div class="panel"><div class="panel-title">📊 各类别月度对比</div><div class="tbl-wrap">';
  h += '<table><thead><tr><th>类别</th>';
  mks.forEach(function(mk){{ h += '<th>' + mk + '合计</th>'; }});
  for (var ci = 1; ci < mks.length; ci++) {{ h += '<th>' + mks[ci] + 'vs' + mks[ci-1] + '</th>'; }}
  h += '<th>趋势</th></tr></thead><tbody>';
  var cats = ['优惠到期', '存量变更', '拆机销户', '纯新套餐', '存量加装'];
  cats.forEach(function(c) {{
    var colVals = mks.map(function(mk) {{ return (MONTH_DATA[mk].catStats[c] || {{sum:0}}).sum; }});
    var trend = '';
    var allUp = colVals.every(function(v,i){{return i===0||v>colVals[i-1];}});
    var allDown = colVals.every(function(v,i){{return i===0||v<colVals[i-1];}});
    if (allUp) trend = '📈 持续上升';
    else if (allDown) trend = '📉 持续下降';
    else if (colVals[0] < colVals[colVals.length-1]) trend = '↘️ 波动上行';
    else if (colVals[0] > colVals[colVals.length-1]) trend = '↗️ 波动下行';
    else trend = '➡️ 波动';
    h += '<tr><td class="tl" style="font-weight:500;">' + c + '</td>';
    mks.forEach(function(mk,i) {{
      var v = colVals[i];
      h += '<td class="tr ' + (v >= 0 ? 'p' : 'n') + '">' + (v >= 0 ? '+' : '') + v.toFixed(2) + '</td>';
    }});
    for (var ci = 1; ci < colVals.length; ci++) {{
      var diff = colVals[ci] - colVals[ci-1];
      h += '<td class="tr ' + (diff >= 0 ? 'p' : 'n') + '">' + (diff >= 0 ? '+' : '') + diff.toFixed(2) + '</td>';
    }}
    h += '<td>' + trend + '</td></tr>';
  }});
  h += '<tr style="font-weight:700;background:#f5f7ff;"><td class="tl">净增合计</td>';
  sums.forEach(function(v,i) {{
    h += '<td class="tr ' + (v >= 0 ? 'p' : 'n') + '">' + (v >= 0 ? '+' : '') + v.toFixed(2) + '</td>';
  }});
  chgPairs.forEach(function(p) {{
    h += '<td class="tr ' + (p.chg >= 0 ? 'p' : 'n') + '">' + (p.chg >= 0 ? '+' : '') + p.chg.toFixed(2) + '</td>';
  }});
  h += '<td></td></tr>';
  h += '</tbody></table></div></div>';
  
  // Key insight box
  h += '<div class="panel" style="background:#f5f7ff;border-left:4px solid #3949ab;padding:14px 18px;">';
  h += '<div class="panel-title" style="border:none;margin:0;">📌 月度环比核心判断</div>';
  var insight = '';
  chgPairs.forEach(function(pair) {{
    var direction = pair.chg >= 0 ? '增长' : '下降';
    insight += pair.to + '较' + pair.from + direction;
    if (pair.chg >= 0) insight += '（净增表现改善）';
    else insight += '（净增表现下滑）';
    insight += pair.pct !== 'N/A' ? '，变化幅度' + pair.pct + '%' : '';
    insight += '。';
  }});
  var bestDayIdx = dayAvgs.indexOf(Math.max.apply(null, dayAvgs));
  insight += '日均表现最好的月份是' + mks[bestDayIdx] + '（' + dayAvgs[bestDayIdx].toFixed(2) + '/天）。';
  h += '<p style="font-size:13px;color:#37474f;line-height:1.7;margin-top:4px;">' + insight + '</p></div>';

  h += '</div>';  // end hb-overview

  // ---- Panel 2: Daily trend comparison ----
  h += '<div class="hb-panel" id="hbp-hb-daily">';
  h += '<div class="chart-box"><h3>📈 各月净增趋势对比（按自然日对齐）</h3>';
  h += '<div class="chart-wrap" style="height:320px;"><canvas id="hbDailyTrend"></canvas></div></div>';
  h += '<div class="chart-box" style="margin-top:12px;"><h3>📊 累计净增对比</h3>';
  h += '<div class="chart-wrap" style="height:320px;"><canvas id="hbCumTrend"></canvas></div></div>';
  h += '<div class="g3" style="margin-top:12px;">';
  for (var mi = 0; mi < mks.length; mi++) {{
    var mk = mks[mi];
    var mkDates = MONTH_DATA[mk].dateList;
    var mkDt = MONTH_DATA[mk].dailyTotals;
    var dayLabel = mkDt.length + '天数据';
    h += '<div class="panel" style="text-align:center;"><div class="panel-title" style="justify-content:center;">' + mk + ' ' + dayLabel + '</div>';
    h += '<div class="chart-wrap" style="height:240px;"><canvas id="hbDaily-' + mk + '"></canvas></div></div>';
  }}
  h += '</div></div>';

  // ---- Panel 3: Grid comparison ----
  h += '<div class="hb-panel" id="hbp-hb-grids">';

  // Get all common grids
  var allGrids = {{}};
  mks.forEach(function(mk) {{
    MONTH_DATA[mk].fullRanking.forEach(function(r) {{ allGrids[r.name] = true; }});
  }});
  var gridNames = Object.keys(allGrids).sort();

  // Build grid comparison data
  var gridComp = [];
  gridNames.forEach(function(name) {{
    var vals = {{}};
    mks.forEach(function(mk) {{
      var found = MONTH_DATA[mk].fullRanking.find(function(r) {{ return r.name === name; }});
      vals[mk] = found ? found.total : null;
    }});
    var chg54 = vals['5月'] !== null && vals['4月'] !== null ? vals['5月'] - vals['4月'] : null;
    var chg65 = vals['6月'] !== null && vals['5月'] !== null ? vals['6月'] - vals['5月'] : null;
    // Time progression: compare periods with same elapsed days
    // 4月 and 5月 are full months, 6月 is MTD. For fairness, compare all as MTD.
    gridComp.push({{ name: name, vals: vals, chg54: chg54, chg65: chg65 }});
  }});

  // Sort by 6月 total
  gridComp.sort(function(a,b){{ return (b.vals['6月'] || 0) - (a.vals['6月'] || 0); }});

  // Summary stats
  var improved54 = gridComp.filter(function(g) {{ return g.chg54 !== null && g.chg54 > 0; }}).length;
  var worsened54 = gridComp.filter(function(g) {{ return g.chg54 !== null && g.chg54 < 0; }}).length;
  var improved65 = gridComp.filter(function(g) {{ return g.chg65 !== null && g.chg65 > 0; }}).length;
  var worsened65 = gridComp.filter(function(g) {{ return g.chg65 !== null && g.chg65 < 0; }}).length;

  h += '<div class="hb-summary">';
  h += '<div class="hb-card"><h3>📊 5月 vs 4月</h3><div class="hl">改善网格 / 恶化网格</div><div class="hv2" style="font-size:18px;"><span class="hb-up">+' + improved54 + '</span> / <span class="hb-down">-' + worsened54 + '</span></div></div>';
  h += '<div class="hb-card"><h3>📊 6月 vs 5月</h3><div class="hl">改善网格 / 恶化网格</div><div class="hv2" style="font-size:18px;"><span class="hb-up">+' + improved65 + '</span> / <span class="hb-down">-' + worsened65 + '</span></div></div>';
  h += '<div class="hb-card"><h3>🏆 持续改善</h3><div class="hl">4-5-6月连续正向增长</div><div class="hv2" style="font-size:18px;color:#2e7d32;">' + gridComp.filter(function(g){{return g.chg54>0 && g.chg65>0}}).length + '</div></div>';
  h += '</div>';

  // Top improvers and decliners
  var topImprovers = gridComp.filter(function(g){{return g.chg65 !== null}}).sort(function(a,b){{return (b.chg65||0)-(a.chg65||0)}}).slice(0,10);
  var topDecliners = gridComp.filter(function(g){{return g.chg65 !== null}}).sort(function(a,b){{return (a.chg65||0)-(b.chg65||0)}}).slice(0,10);

  h += '<div class="g2"><div class="panel"><div class="panel-title" style="color:#2e7d32;">📈 6月环比改善 TOP 10</div><div class="tbl-wrap"><table class="comp-table"><thead><tr><th>#</th><th class="tl">网格名称</th><th>4月</th><th>5月</th><th>6月</th><th>5→6变化</th></tr></thead><tbody>';
  topImprovers.forEach(function(g,i) {{
    h += '<tr><td>' + (i+1) + '</td><td class="tl">' + g.name + '</td>';
    h += '<td class="tr ' + ((g.vals['4月']||0) >= 0 ? 'p' : 'n') + '">' + (g.vals['4月'] !== null ? g.vals['4月'].toFixed(1) : '—') + '</td>';
    h += '<td class="tr ' + ((g.vals['5月']||0) >= 0 ? 'p' : 'n') + '">' + (g.vals['5月'] !== null ? g.vals['5月'].toFixed(1) : '—') + '</td>';
    h += '<td class="tr ' + ((g.vals['6月']||0) >= 0 ? 'p' : 'n') + '">' + (g.vals['6月'] !== null ? g.vals['6月'].toFixed(1) : '—') + '</td>';
    h += '<td class="tr ' + ((g.chg65||0) >= 0 ? 'p' : 'n') + '">' + (g.chg65 >= 0 ? '+' : '') + g.chg65.toFixed(1) + '</td></tr>';
  }});
  h += '</tbody></table></div></div>';

  h += '<div class="panel"><div class="panel-title" style="color:#c62828;">📉 6月环比恶化 TOP 10</div><div class="tbl-wrap"><table class="comp-table"><thead><tr><th>#</th><th class="tl">网格名称</th><th>4月</th><th>5月</th><th>6月</th><th>5→6变化</th></tr></thead><tbody>';
  topDecliners.forEach(function(g,i) {{
    h += '<tr><td>' + (i+1) + '</td><td class="tl">' + g.name + '</td>';
    h += '<td class="tr ' + ((g.vals['4月']||0) >= 0 ? 'p' : 'n') + '">' + (g.vals['4月'] !== null ? g.vals['4月'].toFixed(1) : '—') + '</td>';
    h += '<td class="tr ' + ((g.vals['5月']||0) >= 0 ? 'p' : 'n') + '">' + (g.vals['5月'] !== null ? g.vals['5月'].toFixed(1) : '—') + '</td>';
    h += '<td class="tr ' + ((g.vals['6月']||0) >= 0 ? 'p' : 'n') + '">' + (g.vals['6月'] !== null ? g.vals['6月'].toFixed(1) : '—') + '</td>';
    h += '<td class="tr ' + ((g.chg65||0) >= 0 ? 'p' : 'n') + '">' + (g.chg65 >= 0 ? '+' : '') + g.chg65.toFixed(1) + '</td></tr>';
  }});
  h += '</tbody></table></div></div></div>';

  // All grid comparison table
  h += '<div class="panel"><div class="panel-title">📋 所有网格月度净增对比</div>';
  h += '<div class="scroll-y"><table class="comp-table"><thead><tr><th style="width:40px">#</th><th class="tl">网格名称</th><th>4月</th><th>5月</th><th>6月</th><th>5月vs4月</th><th>6月vs5月</th><th>趋势</th></tr></thead><tbody>';
  gridComp.forEach(function(g,i) {{
    var trend = '';
    if (g.chg54 > 0 && g.chg65 > 0) trend = '📈';
    else if (g.chg54 < 0 && g.chg65 < 0) trend = '📉';
    else if (g.chg54 === null || g.chg65 === null) trend = '—';
    else trend = '🔀';
    h += '<tr><td>' + (i+1) + '</td><td class="tl">' + g.name + '</td>';
    h += '<td class="tr ' + ((g.vals['4月']||0) >= 0 ? 'p' : 'n') + '">' + (g.vals['4月'] !== null ? g.vals['4月'].toFixed(1) : '—') + '</td>';
    h += '<td class="tr ' + ((g.vals['5月']||0) >= 0 ? 'p' : 'n') + '">' + (g.vals['5月'] !== null ? g.vals['5月'].toFixed(1) : '—') + '</td>';
    h += '<td class="tr ' + ((g.vals['6月']||0) >= 0 ? 'p' : 'n') + '">' + (g.vals['6月'] !== null ? g.vals['6月'].toFixed(1) : '—') + '</td>';
    h += '<td class="tr ' + ((g.chg54||0) >= 0 ? 'p' : 'n') + '">' + (g.chg54 !== null ? (g.chg54 >= 0 ? '+' : '') + g.chg54.toFixed(1) : '—') + '</td>';
    h += '<td class="tr ' + ((g.chg65||0) >= 0 ? 'p' : 'n') + '">' + (g.chg65 !== null ? (g.chg65 >= 0 ? '+' : '') + g.chg65.toFixed(1) : '—') + '</td>';
    h += '<td>' + trend + '</td></tr>';
  }});
  h += '</tbody></table></div></div>';
  h += '</div>';  // end hb-grids

  // ---- Panel 4: Rank changes ----
  h += '<div class="hb-panel" id="hbp-hb-detail">';
  
  // Get rank for each month
  var rankMap = {{}};
  mks.forEach(function(mk) {{
    MONTH_DATA[mk].fullRanking.forEach(function(r) {{
      if (!rankMap[r.name]) rankMap[r.name] = {{}};
      rankMap[r.name][mk] = r.rank;
    }});
  }});

  var rankChanges = [];
  gridNames.forEach(function(name) {{
    var r4 = rankMap[name]['4月'] || 999;
    var r5 = rankMap[name]['5月'] || 999;
    var r6 = rankMap[name]['6月'] || 999;
    var chgRank54 = r5 - r4;
    var chgRank65 = r6 - r5;
    rankChanges.push({{ name: name, r4: r4, r5: r5, r6: r6, chg54: chgRank54, chg65: chgRank65 }});
  }});

  // Top rank climbers this month
  var climbers = rankChanges.filter(function(g){{return g.chg65 < 0 && g.chg65 !== null}}).sort(function(a,b){{return a.chg65-b.chg65}}).slice(0,10);
  var fallers = rankChanges.filter(function(g){{return g.chg65 > 0 && g.chg65 !== null}}).sort(function(a,b){{return b.chg65-a.chg65}}).slice(0,10);

  h += '<div class="g2"><div class="panel"><div class="panel-title" style="color:#2e7d32;">🏆 排名上升 TOP 10（vs 5月）</div><div class="tbl-wrap"><table class="comp-table"><thead><tr><th>#</th><th class="tl">网格名称</th><th>4月排名</th><th>5月排名</th><th>6月排名</th><th>排名变化</th></tr></thead><tbody>';
  climbers.forEach(function(g,i) {{
    h += '<tr><td>' + (i+1) + '</td><td class="tl">' + g.name + '</td>';
    h += '<td class="tr">' + (g.r4 < 999 ? g.r4 : '—') + '</td>';
    h += '<td class="tr">' + (g.r5 < 999 ? g.r5 : '—') + '</td>';
    h += '<td class="tr" style="font-weight:600;">' + (g.r6 < 999 ? g.r6 : '—') + '</td>';
    h += '<td class="tr hb-up">↑' + Math.abs(g.chg65) + '</td></tr>';
  }});
  h += '</tbody></table></div></div>';

  h += '<div class="panel"><div class="panel-title" style="color:#c62828;">📉 排名下降 TOP 10（vs 5月）</div><div class="tbl-wrap"><table class="comp-table"><thead><tr><th>#</th><th class="tl">网格名称</th><th>4月排名</th><th>5月排名</th><th>6月排名</th><th>排名变化</th></tr></thead><tbody>';
  fallers.forEach(function(g,i) {{
    h += '<tr><td>' + (i+1) + '</td><td class="tl">' + g.name + '</td>';
    h += '<td class="tr">' + (g.r4 < 999 ? g.r4 : '—') + '</td>';
    h += '<td class="tr">' + (g.r5 < 999 ? g.r5 : '—') + '</td>';
    h += '<td class="tr" style="font-weight:600;">' + (g.r6 < 999 ? g.r6 : '—') + '</td>';
    h += '<td class="tr hb-down">↓' + Math.abs(g.chg65) + '</td></tr>';
  }});
  h += '</tbody></table></div></div></div>';

  h += '</div>';  // end hb-detail

  document.getElementById('mp-环比分析').innerHTML = h;

  // Boot HB charts
  setTimeout(function() {{ bootHBCharts(); }}, 100);
}}

function bootHBCharts() {{
  var mks = ['1月','2月','3月','4月','5月','6月','7月'];
  var colors = ['#f9a825','#1a237e','#c62828','#2e7d32','#e65100','#6a1b9a','#00838f'];
  var monthPrefix = {{'1月':'01','2月':'02','3月':'03','4月':'04','5月':'05','6月':'06','7月':'07'}};
  
  {{ // Daily trend comparison - align by day index (1-16 for common days)
    var canv = document.getElementById('hbDailyTrend');
    if (canv) {{
      var ctx = canv.getContext('2d');
      if (charts['hb-trend']) charts['hb-trend'].destroy();
      var maxDays = 31;  // May has most days
      var dayLabels = [];
      for (var d = 1; d <= maxDays; d++) dayLabels.push('D' + d);
      
      var datasets = [];
      mks.forEach(function(mk, mi) {{
        var dt = MONTH_DATA[mk].dailyTotals;
        var data = [];
        for (var d = 1; d <= maxDays; d++) {{
          var dateStr = '2026' + (monthPrefix[mk] || '06') + (d < 10 ? '0' : '') + d;
          var found = dt.find(function(row) {{ return row.date === dateStr; }});
          data.push(found ? found.total : null);
        }}
        datasets.push({{
          label: mk, data: data, borderColor: colors[mi],
          backgroundColor: colors[mi] + '22',
          fill: false, tension: 0.3, pointRadius: 3, spanGaps: true,
          borderWidth: 2
        }});
      }});
      charts['hb-trend'] = new Chart(ctx, {{
        type: 'line',
        data: {{ labels: dayLabels, datasets: datasets }},
        options: {{
          responsive: true, maintainAspectRatio: false,
          plugins: {{ legend: {{ position: 'top', labels: {{ boxWidth: 12, font: {{ size: 11 }} }} }} }},
          scales: {{ y: {{ grid: {{ color: 'rgba(0,0,0,0.05)' }} }}, x: {{ ticks: {{ maxRotation: 45, font: {{ size: 9 }} }} }} }}
        }}
      }});
    }}
  }}

  {{ // Cumulative chart
    var canv = document.getElementById('hbCumTrend');
    if (canv) {{
      var ctx = canv.getContext('2d');
      if (charts['hb-cum']) charts['hb-cum'].destroy();
      var maxDays = 31;
      var dayLabels = [];
      for (var d = 1; d <= maxDays; d++) dayLabels.push('D' + d);
      
      var datasets = [];
      mks.forEach(function(mk, mi) {{
        var dt = MONTH_DATA[mk].dailyTotals;
        var data = [];
        var cum = 0;
        for (var d = 1; d <= maxDays; d++) {{
          var dateStr = '2026' + (monthPrefix[mk] || '06') + (d < 10 ? '0' : '') + d;
          var found = dt.find(function(row) {{ return row.date === dateStr; }});
          if (found) cum += found.total;
          data.push(Math.round(cum * 100) / 100);
        }}
        datasets.push({{
          label: mk, data: data, borderColor: colors[mi],
          backgroundColor: colors[mi] + '22',
          fill: true, tension: 0.3, pointRadius: 2, spanGaps: true,
          borderWidth: 2
        }});
      }});
      charts['hb-cum'] = new Chart(ctx, {{
        type: 'line',
        data: {{ labels: dayLabels, datasets: datasets }},
        options: {{
          responsive: true, maintainAspectRatio: false,
          plugins: {{ legend: {{ position: 'top', labels: {{ boxWidth: 12, font: {{ size: 11 }} }} }} }},
          scales: {{ y: {{ grid: {{ color: 'rgba(0,0,0,0.05)' }} }}, x: {{ ticks: {{ maxRotation: 45, font: {{ size: 9 }} }} }} }}
        }}
      }});
    }}
  }}

  {{ // Individual month daily charts
    mks.forEach(function(mk) {{
      var canv = document.getElementById('hbDaily-' + mk);
      if (!canv) return;
      var ctx = canv.getContext('2d');
      if (charts['hb-ind-' + mk]) charts['hb-ind-' + mk].destroy();
      var dt = MONTH_DATA[mk].dailyTotals;
      var labels = dt.map(function(d) {{ return d.date.slice(4,6) + '-' + d.date.slice(6,8); }});
      var posData = dt.map(function(d) {{ return Math.max(0, d.total); }});
      var negData = dt.map(function(d) {{ return Math.min(0, d.total); }});
      charts['hb-ind-' + mk] = new Chart(ctx, {{
        type: 'bar',
        data: {{
          labels: labels,
          datasets: [
            {{ label: '正净增', data: posData, backgroundColor: 'rgba(46,125,50,0.7)', borderRadius: 2 }},
            {{ label: '负净增', data: negData, backgroundColor: 'rgba(198,40,40,0.7)', borderRadius: 2 }}
          ]
        }},
        options: {{
          responsive: true, maintainAspectRatio: false,
          plugins: {{ legend: {{ display: false }} }},
          scales: {{ y: {{ grid: {{ color: 'rgba(0,0,0,0.05)' }} }}, x: {{ ticks: {{ maxRotation: 45, font: {{ size: 9 }} }} }} }}
        }}
      }});
    }});
  }}
}}

// ---- Init ----
function init() {{
  // Build month tabs
  var mtabHTML = '';
  var monthNames = ['1月','2月','3月','4月','5月','6月','7月','📊 环比分析','💡 优化建议'];
  monthNames.forEach(function(mn, i) {{
    mtabHTML += '<button class="mtab-btn' + (i===0?' active':'') + '" onclick="switchMonth(' + i + ')">' + mn + '</button>';
  }});
  document.getElementById('monthTabs').innerHTML = mtabHTML;

  // Init each month
  renderMonth('1月');
  renderMonth('2月');
  renderMonth('3月');
  renderMonth('4月');
  renderMonth('5月');
  renderMonth('6月');
  renderMonth('7月');
  renderMonthComparison();

  // Update topbar
  var d4 = MONTH_DATA['4月'];
  document.getElementById('topGridCount').textContent = '🏘️ ' + d4.totalGrids + ' 个网格';
  document.getElementById('topDateRange').textContent = '📆 ' + (d4.dateList.length + MONTH_DATA['5月'].dateList.length + MONTH_DATA['6月'].dateList.length) + ' 天数据';
}}

var currentMonth = 0;
function switchMonth(idx) {{
  currentMonth = idx;
  var panels = document.querySelectorAll('.mtab-panel');
  panels.forEach(function(p) {{ p.classList.remove('active'); }});
  var btns = document.querySelectorAll('.mtab-btn');
  btns.forEach(function(b) {{ b.classList.remove('active'); }});
  btns[idx].classList.add('active');
  
  var panelId = 'mp-' + (['1月','2月','3月','4月','5月','6月','7月','环比分析','优化建议'])[idx];
  document.getElementById(panelId).classList.add('active');
  
  // Re-render charts (fix resize issue)
  if (idx < 7) {{
    var mk = ['1月','2月','3月','4月','5月','6月','7月'][idx];
    setTimeout(function() {{ bootMonthCharts(mk); }}, 50);
  }}
}}

// Sub-tab handlers
document.addEventListener('click', function(e) {{
  var target = e.target;
  if (target.classList.contains('sub-tab-btn')) {{
    var sub = target.getAttribute('data-sub');
    var mk = sub.split('-')[0];
    var panelId = sub.split('-')[1];
    // Find parent mp
    var container = target.closest('.mtab-panel');
    if (!container) return;
    var btns = container.querySelectorAll('.sub-tab-btn');
    btns.forEach(function(b) {{ b.classList.remove('active'); }});
    target.classList.add('active');
    var panels = container.querySelectorAll('.sub-tab-panel');
    panels.forEach(function(p) {{ p.classList.remove('active'); }});
    document.getElementById('st-' + sub).classList.add('active');
  }}
  if (target.classList.contains('hb-tab-btn')) {{
    var hb = target.getAttribute('data-hb');
    var container = document.getElementById('mp-环比分析');
    var btns = container.querySelectorAll('.hb-tab-btn');
    btns.forEach(function(b) {{ b.classList.remove('active'); }});
    target.classList.add('active');
    var panels = container.querySelectorAll('.hb-panel');
    panels.forEach(function(p) {{ p.classList.remove('active'); }});
    document.getElementById('hbp-' + hb).classList.add('active');
  }}
}});

// Resize handler
window.addEventListener('resize', function() {{
  var idx = currentMonth;
  if (idx < 7) {{
    var mk = ['1月','2月','3月','4月','5月','6月','7月'][idx];
    setTimeout(function() {{ bootMonthCharts(mk); }}, 200);
    setTimeout(function() {{ bootHBCharts(); }}, 200);
  }}
}});

init();
</script>
'''

html += f'''
<div class="footer">网格积分数据看板 · 2026年1月-7月 · 数据每日更新</div>
</body>
</html>'''

OUTPUT_DIR = "/Users/mr.g/Documents/Codex/Workspace/projects/project-001-grid-data-dashboard/docs"
os.makedirs(OUTPUT_DIR, exist_ok=True)
outpath = os.path.join(OUTPUT_DIR, "dashboard.html")
with open(outpath, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"\n✅ 看板已生成: {outpath}")
print(f"文件大小: {os.path.getsize(outpath) / 1024:.0f} KB")



