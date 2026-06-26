#!/usr/bin/env python3
"""
网格积分每日分析脚本
用法：python3 src/daily_report.py <Excel文件路径> [--date 统计日期]

数据格式要求：
  必须包含列：网格单元名称, 优惠到期, 存量变更, 拆机销户, 纯新套餐, 存量加装
  可选包含列：统计日期, 合计
"""

import sys, os, json
import openpyxl

def to_num(v):
    if v is None: return None
    s = str(v).strip().lower()
    if s in ('<null>', 'none', 'null', ''): return None
    try: return float(v)
    except: return None

def analyze(filepath, report_date=None):
    """主分析函数，返回分析结果字典"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Read header to find columns
    header = [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    
    # Known category columns
    cat_names = ['优惠到期', '存量变更', '拆机销户', '纯新套餐', '存量加装']
    col_map = {}  # col index (1-based) -> column name
    
    date_col = None
    name_col = None
    
    for i, h in enumerate(header, 1):
        if h == '网格单元名称':
            name_col = i
        elif h in cat_names:
            col_map[i] = h
        elif h == '统计日期':
            date_col = i

    if name_col is None or len(col_map) == 0:
        wb.close()
        raise ValueError(f"无法识别列结构: {header}")

    # Read rows
    records = []
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, name_col).value or '').strip()
        if not name:
            continue
        
        vals = {}
        for ci, cn in col_map.items():
            vals[cn] = to_num(ws.cell(r, ci).value)
        
        row_date = None
        if date_col:
            dv = ws.cell(r, date_col).value
            if dv:
                row_date = str(dv).strip()
        
        total = sum(v for v in vals.values() if v is not None)
        records.append({
            'name': name, 'vals': vals, 'total': round(total, 2),
            'date': row_date or report_date
        })
    
    wb.close()

    if not records:
        return {'error': '无有效数据'}

    # Use report date from filename or data
    if not report_date and records[0]['date']:
        report_date = records[0]['date']
    if not report_date:
        report_date = "未知日期"

    # Rank by total
    records.sort(key=lambda x: x['total'], reverse=True)

    # Stats
    cats = list(col_map.values())
    result = {
        'report_date': report_date,
        'total_grids': len(records),
        'avg_net': round(sum(r['total'] for r in records)/len(records), 2),
        'sum_net': round(sum(r['total'] for r in records), 2),
        'pos_count': sum(1 for r in records if r['total'] > 0),
        'neg_count': sum(1 for r in records if r['total'] < 0),
        'zero_count': sum(1 for r in records if r['total'] == 0),
        'top10': [(r['name'], r['total']) for r in records[:10]],
        'bot5': [(r['name'], r['total']) for r in records[-5:]],
        'categories': {},
        'ranking': [(r['name'], r['total'], {k: r['vals'].get(k, None) for k in cats}) for r in records],
        'categories_order': cats,
    }

    for cat in cats:
        vals = [r['vals'][cat] for r in records if r['vals'][cat] is not None]
        pos = [v for v in vals if v > 0]
        neg = [v for v in vals if v < 0]
        top_p = sorted(
            [(r['name'], r['vals'][cat]) for r in records if r['vals'][cat] is not None and r['vals'][cat] > 0],
            key=lambda x: -x[1])[:5]
        top_n = sorted(
            [(r['name'], r['vals'][cat]) for r in records if r['vals'][cat] is not None and r['vals'][cat] < 0],
            key=lambda x: x[1])[:5]
        result['categories'][cat] = {
            'count': len(vals), 'sum': round(sum(vals), 2),
            'avg': round(sum(vals)/len(vals), 2) if vals else 0,
            'max': round(max(vals), 2) if vals else 0,
            'min': round(min(vals), 2) if vals else 0,
            'pos': len(pos), 'neg': len(neg),
            'rate': f"{len(vals)/len(records)*100:.1f}%",
            'top_pos': top_p, 'top_neg': top_n,
        }

    return result


def print_report_text(result):
    """输出文字版报告"""
    d = result
    print(f"\n{'='*60}")
    print(f"  网格积分分析报告 — {d['report_date']}")
    print(f"{'='*60}")
    print(f"\n网格总数: {d['total_grids']} | 正净增: {d['pos_count']} | 负净增: {d['neg_count']}")
    print(f"净增合计: {d['sum_net']:.2f} | 平均值: {d['avg_net']:.2f}")
    
    print(f"\n--- TOP 5 ---")
    for i, (n, t) in enumerate(d['top10'][:5], 1):
        tag = "+" if t >= 0 else ""
        print(f"  {i}. {n:<24} {tag}{t:>8.2f}")
    
    print(f"\n--- BOTTOM 5 ---")
    for i, (n, t) in enumerate(d['bot5'], 1):
        print(f"  {i}. {n:<24} {t:>8.2f}")
    
    print(f"\n--- 各类型统计 ---")
    for cat in d['categories_order']:
        s = d['categories'][cat]
        print(f"  {cat}: 合计={s['sum']:>8.2f} | 有数据={s['rate']} | 正={s['pos']} 负={s['neg']}")
    
    print(f"\n报告位置: docs/{d['report_date']}_分析报告.html")
    print(f"{'='*60}")


def gen_html(result):
    """生成报告HTML"""
    d = result
    cats = d['categories_order']
    ranking_rows = ""
    for rank, (name, total, vals) in enumerate(d['ranking'], 1):
        cls = "pos" if total >= 0 else "neg"
        cells = f"<td class=\"num-col\">{rank}</td><td>{name}</td><td class=\"num-col {cls}\">{total:>8.2f}</td>"
        for cat in cats:
            v = vals.get(cat)
            if v is not None:
                c = "pos" if v > 0 else ("neg" if v < 0 else "")
                cells += f"<td class=\"num-col {c}\">{v:>8.2f}</td>"
            else:
                cells += "<td class=\"num-col\" style=\"opacity:.3\">—</td>"
        ranking_rows += f"<tr>{cells}</tr>"

    cat_rows = ""
    for cat in cats:
        s = d['categories'][cat]
        cls = "pos" if s['sum'] >= 0 else "neg"
        cat_rows += f"""<tr><td>{cat}</td><td class="num-col {cls}">{s['sum']:>+.2f}</td>
        <td class="num-col">{s['avg']:.2f}</td><td class="num-col">{s['max']:.2f}</td>
        <td class="num-col">{s['min']:.2f}</td><td class="num-col">{s['rate']}</td>
        <td class="num-col">{s['pos']}</td><td class="num-col">{s['neg']}</td></tr>"""

    top10_rows = ""
    for rank, (name, total) in enumerate(d['top10'], 1):
        cls = "pos" if total >= 0 else "neg"
        top10_rows += f"<tr><td class=\"num-col\">{rank}</td><td>{name}</td><td class=\"num-col {cls}\">{total:>8.2f}</td></tr>"

    bot5_rows = ""
    for rank_offset, (name, total) in enumerate(d['bot5'], 1):
        rank = d['total_grids'] - 5 + rank_offset
        bot5_rows += f"<tr><td class=\"num-col\">{rank}</td><td>{name}</td><td class=\"num-col neg\">{total:>8.2f}</td></tr>"

    html = f"""<!DOCTYPE html>
<html lang="zh-CN"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>网格积分分析报告 {d['report_date']}</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+SC:wght@400;500;700&display=swap');
  *{{margin:0;padding:0;box-sizing:border-box}}
  body{{font-family:'Noto Sans SC',sans-serif;background:#f0f2f5;padding:40px 20px;color:#1a1a2e}}
  .page{{max-width:960px;margin:0 auto;background:#fff;border-radius:12px;box-shadow:0 8px 40px rgba(0,0,0,0.1);overflow:hidden}}
  .header{{background:linear-gradient(135deg,#1a237e,#3949ab);color:#fff;padding:36px 50px 28px;position:relative}}
  .header::after{{content:'';position:absolute;bottom:0;left:0;right:0;height:6px;background:linear-gradient(90deg,#ffd54f,#ffb300,#ffd54f)}}
  .header .date-tag{{font-size:13px;opacity:.7;margin-bottom:6px}}
  .header h1{{font-size:26px;letter-spacing:2px}}
  .body{{padding:36px 50px 44px}}
  .summary-row{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:28px}}
  .summary-card{{padding:16px;border-radius:10px;text-align:center}}
  .summary-card .num{{font-size:24px;font-weight:700}}
  .summary-card .label{{font-size:12px;opacity:.7;margin-top:4px}}
  .sc-blue{{background:#e3f2fd;color:#1565c0}}
  .sc-green{{background:#e8f5e9;color:#2e7d32}}
  .sc-red{{background:#fce4ec;color:#c62828}}
  .sc-orange{{background:#fff3e0;color:#e65100}}
  .section{{margin-bottom:30px}}
  .section-title{{font-size:17px;font-weight:700;color:#1a237e;padding-bottom:8px;border-bottom:2px solid #e8eaf6;margin-bottom:14px}}
  table{{width:100%;border-collapse:collapse;font-size:13px;margin-bottom:10px}}
  th{{background:#e8eaf6;color:#1a237e;font-weight:600;padding:8px 10px;text-align:left;font-size:12px}}
  td{{padding:6px 10px;border-bottom:1px solid #f0f0f0;font-size:13px}}
  tr:hover td{{background:#fafbff}}
  .num-col{{text-align:right;font-variant-numeric:tabular-nums;font-family:'Menlo',monospace;font-size:12px}}
  .pos{{color:#2e7d32}}
  .neg{{color:#c62828}}
  .grid-2{{display:grid;grid-template-columns:1fr 1fr;gap:14px}}
  .stat-card{{border:1px solid #e0e0e0;border-radius:8px;padding:14px}}
  .stat-card h4{{font-size:13px;color:#1a237e;margin-bottom:6px}}
  @media(max-width:600px){{body{{padding:12px 8px}}.header{{padding:24px 16px 20px}}.header h1{{font-size:20px}}.body{{padding:20px 12px 30px}}.summary-row{{grid-template-columns:repeat(2,1fr)}}
  .grid-2{{grid-template-columns:1fr}}}}
</style></head><body>
<div class="page">
  <div class="header">
    <div class="date-tag">📅 {d['report_date']}</div>
    <h1>网格积分发展分析报告</h1>
  </div>
  <div class="body">
    <div class="section">
      <div class="section-title">📊 整体概览</div>
      <div class="summary-row">
        <div class="summary-card sc-blue"><div class="num">{d['total_grids']}</div><div class="label">网格总数</div></div>
        <div class="summary-card sc-green"><div class="num" style="color:#2e7d32;">{d['pos_count']}</div><div class="label">正净增网格</div></div>
        <div class="summary-card sc-red"><div class="num" style="color:#c62828;">{d['neg_count']}</div><div class="label">负净增网格</div></div>
        <div class="summary-card sc-orange"><div class="num" style="color:#e65100;">{d['avg_net']:.2f}</div><div class="label">平均净增</div></div>
      </div>
    </div>
    <div class="section">
      <div class="section-title">🏆 TOP 10</div>
      <table><tr><th>排名</th><th>网格名称</th><th>净增积分</th></tr>{top10_rows}</table>
    </div>
    <div class="section">
      <div class="section-title">⚠️ BOTTOM 5</div>
      <table><tr><th>排名</th><th>网格名称</th><th>净增积分</th></tr>{bot5_rows}</table>
    </div>
    <div class="section">
      <div class="section-title">📈 各类型统计</div>
      <table><tr><th>类型</th><th>合计</th><th>平均值</th><th>最大值</th><th>最小值</th><th>有数据率</th><th>正数</th><th>负数</th></tr>{cat_rows}</table>
    </div>
    <div class="section">
      <div class="section-title">📋 全部网格排名</div>
      <table>
        <tr><th>排名</th><th>网格名称</th><th>净增积分</th>{"".join(f'<th>{c}</th>' for c in cats)}</tr>
        {ranking_rows}
      </table>
    </div>
  </div>
</div>
<script src="https://cdn.jsdelivr.net/npm/html2canvas@1.4.1/dist/html2canvas.min.js"></script>
<style>
  @keyframes spin{{to{{transform:rotate(360deg)}}}}
  .toolbar{{max-width:960px;margin:16px auto 0;display:flex;justify-content:flex-end}}
  .btn-save{{display:inline-flex;align-items:center;gap:6px;padding:10px 22px;background:linear-gradient(135deg,#1a237e,#3949ab);color:#fff;border:none;border-radius:8px;font-size:14px;font-weight:600;cursor:pointer;font-family:inherit;box-shadow:0 2px 8px rgba(26,35,126,0.25)}}
  .btn-save.loading .label-text{{display:none}}
  .btn-save.loading .spinner{{display:inline-block}}
  .spinner{{display:none;width:16px;height:16px;border:2px solid rgba(255,255,255,0.3);border-top-color:#fff;border-radius:50%;animation:spin .6s linear infinite}}
  @media print{{.toolbar{{display:none}}}}
</style>
<div class="toolbar">
  <button class="btn-save" id="saveBtn" onclick="saveAsImage()">
    <span class="spinner"></span><span class="label-text">📷 保存为长图</span>
  </button>
</div>
<script>
async function saveAsImage(){{const btn=document.getElementById('saveBtn');const area=document.querySelector('.page');
btn.classList.add('loading');btn.disabled=true;
try{{const c=await html2canvas(area,{{scale:2,useCORS:true,backgroundColor:'#ffffff',logging:false,width:area.scrollWidth,height:area.scrollHeight,windowWidth:area.scrollWidth,windowHeight:area.scrollHeight}});
const l=document.createElement('a');l.download='{d["report_date"]}_网格分析报告.png';l.href=c.toDataURL('image/png');l.click()}}
catch(e){{alert('生成失败：'+e.message)}}
finally{{btn.classList.remove('loading');btn.disabled=false}}}}
</script>
</body></html>"""
    return html


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("用法: python3 src/daily_report.py <Excel文件路径> [--date 统计日期]")
        sys.exit(1)

    filepath = sys.argv[1]
    date_arg = None
    if '--date' in sys.argv:
        idx = sys.argv.index('--date')
        if idx + 1 < len(sys.argv):
            date_arg = sys.argv[idx + 1]

    result = analyze(filepath, date_arg)
    if 'error' in result:
        print(f"错误: {result['error']}")
        sys.exit(1)

    print_report_text(result)
    
    # Save HTML report
    report_date = result['report_date']
    html = gen_html(result)
    outdir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'docs')
    os.makedirs(outdir, exist_ok=True)
    outpath = os.path.join(outdir, f"{report_date}_分析报告.html")
    with open(outpath, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"\n✅ 报告已生成: {outpath}")
