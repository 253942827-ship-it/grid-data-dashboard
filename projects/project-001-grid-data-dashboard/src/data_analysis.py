import openpyxl
import csv

fpath = "assets/网格单元运营数据202605.xlsx"
wb = openpyxl.load_workbook(fpath)
ws = wb.active

headers = [str(ws.cell(1, c).value) for c in range(1, 8)]
categories = headers[1:6]  # 优惠到期, 存量变更, 拆机销户, 纯新套餐, 存量加装

# Read all data
grids = []
for r in range(2, ws.max_row + 1):
    name = str(ws.cell(r, 1).value or '').strip()
    if not name:
        continue
    vals = {}
    for c_idx, cat in enumerate(categories, 2):
        v = ws.cell(r, c_idx).value
        if v is None:
            vals[cat] = None
        else:
            try:
                vals[cat] = float(v)
            except (ValueError, TypeError):
                vals[cat] = None
    grids.append({'name': name, 'values': vals})

wb.close()

# ============================================================
# 1. 净增积分统计
# ============================================================
print("=" * 75)
print("一、净增积分统计（各网格汇总积分，按高低排序）")
print("=" * 75)

grid_scores = []
for g in grids:
    total = sum(v for v in g['values'].values() if v is not None)
    grid_scores.append((g['name'], total, g['values']))

grid_scores.sort(key=lambda x: x[1], reverse=True)

print(f"\n{'排名':<4} {'网格名称':<26} {'净增积分':>10} {'优惠到期':>8} {'存量变更':>8} {'拆机销户':>8} {'纯新套餐':>8} {'存量加装':>8}")
print("-" * 95)
for rank, (name, total, vals) in enumerate(grid_scores, 1):
    v1 = f"{vals['优惠到期'] or 0:>8.2f}"
    v2 = f"{vals['存量变更'] or 0:>8.2f}"
    v3 = f"{vals['拆机销户'] or 0:>8.2f}"
    v4 = f"{vals['纯新套餐'] or 0:>8.2f}"
    v5 = f"{vals['存量加装'] or 0:>8.2f}"
    print(f"{rank:<4} {name:<26} {total:>8.2f}  {v1} {v2} {v3} {v4} {v5}")

total_avg = sum(s[1] for s in grid_scores) / len(grid_scores)
positive_count = sum(1 for s in grid_scores if s[1] > 0)
negative_count = sum(1 for s in grid_scores if s[1] < 0)
zero_count = sum(1 for s in grid_scores if s[1] == 0)
print(f"\n--- 汇总 ---")
print(f"网格总数：{len(grid_scores)}")
print(f"综合积分平均值：{total_avg:.2f}")
print(f"最高积分：{grid_scores[0][1]:.2f}（{grid_scores[0][0]}）")
print(f"最低积分：{grid_scores[-1][1]:.2f}（{grid_scores[-1][0]}）")
print(f"正积分网格：{positive_count} | 负积分网格：{negative_count} | 零积分网格：{zero_count}")

print(f"\n  TOP 5：")
for name, total, _ in grid_scores[:5]:
    print(f"    {name:<26} {total:>8.2f}")
print(f"  BOTTOM 5：")
for name, total, _ in grid_scores[-5:]:
    print(f"    {name:<26} {total:>8.2f}")

# ============================================================
# 2. 总量统计
# ============================================================
print("\n" + "=" * 75)
print("二、总量统计（各类型汇总数据）")
print("=" * 75)

total_stats = {}
for cat in categories:
    vals = [g['values'][cat] for g in grids if g['values'][cat] is not None]
    if vals:
        total_stats[cat] = {
            '总和': sum(vals), '平均值': sum(vals)/len(vals),
            '最大值': max(vals), '最小值': min(vals),
            '有数据': len(vals), '正数': sum(1 for v in vals if v > 0),
            '负数': sum(1 for v in vals if v < 0), '零值': sum(1 for v in vals if v == 0),
        }

print(f"\n{'类型':<12} {'总和':>10} {'平均值':>10} {'最大值':>10} {'最小值':>10} {'有数据':>6} {'正数':>6} {'负数':>6}")
print("-" * 78)
for cat in categories:
    s = total_stats[cat]
    print(f"{cat:<12} {s['总和']:>10.2f} {s['平均值']:>10.2f} {s['最大值']:>10.2f} {s['最小值']:>10.2f} {s['有数据']:>6} {s['正数']:>6} {s['负数']:>6}")

grand_total = sum(s['总和'] for s in total_stats.values())
print(f"\n所有类型合计：{grand_total:.2f}")

# ============================================================
# 3. 每个类型的发生情况统计
# ============================================================
print("\n" + "=" * 75)
print("三、每个类型的发生情况统计")
print("=" * 75)

for cat in categories:
    vals = [(g['name'], g['values'][cat]) for g in grids]
    has_data = [(n, v) for n, v in vals if v is not None]
    no_data = [n for n, v in vals if v is None]
    positive = [(n, v) for n, v in has_data if v > 0]
    negative = [(n, v) for n, v in has_data if v < 0]
    zero = [(n, v) for n, v in has_data if v == 0]

    print(f"\n--- {cat} ---")
    print(f"  有数据：{len(has_data)}/{len(grids)}（{len(has_data)/len(grids)*100:.1f}%）| 正数：{len(positive)} | 负数：{len(negative)} | 零值：{len(zero)} | 无数据：{len(no_data)}")
    if positive:
        positive.sort(key=lambda x: x[1], reverse=True)
        print(f"  正数 TOP 3：{', '.join(f'{n}({v:.1f})' for n,v in positive[:3])}")
    if negative:
        negative.sort(key=lambda x: x[1])
        print(f"  负数 TOP 3：{', '.join(f'{n}({v:.1f})' for n,v in negative[:3])}")

# ============================================================
# 4. 保存CSV
# ============================================================
with open('docs/数据分析结果.csv', 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.writer(f)
    writer.writerow(['排名', '网格名称', '净增积分', '优惠到期', '存量变更', '拆机销户', '纯新套餐', '存量加装'])
    for rank, (name, total, vals) in enumerate(grid_scores, 1):
        writer.writerow([rank, name, round(total, 2)] + [
            round(vals[cat], 2) if vals[cat] is not None else '' for cat in categories
        ])

print(f"\n{'='*75}")
print("详细数据已保存到 docs/数据分析结果.csv")
print("=" * 75)
